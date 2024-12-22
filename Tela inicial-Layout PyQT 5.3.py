Tela inicial-Layout PyQT 5.3.py


import os
import sys
import json
import shutil
from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *
import fitz
from PySide6.QtGui import QImage, QPixmap
from PySide6.QtWebEngineWidgets import QWebEngineView
from PySide6.QtWidgets import (QDialog, QTableWidget, QTableWidgetItem, 
                            QDialogButtonBox, QVBoxLayout, QMessageBox)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from PySide6.QtWidgets import QFileDialog
from PySide6.QtCore import QTimer
from PySide6.QtCore import Signal
from PySide6.QtWebEngineCore import (
    QWebEngineSettings, 
    QWebEnginePage, 
    QWebEngineProfile
)
import webbrowser
import win32com.client
from PySide6.QtWebEngineWidgets import QWebEngineView
from PySide6.QtWidgets import QLineEdit, QPushButton, QHBoxLayout

class ClickableLabel(QLabel):
    """Label clicável para o thumbnail"""
    clicked = Signal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCursor(Qt.PointingHandCursor)
    
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.clicked.emit()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Visualização de Pagamentos")
        self.setGeometry(0, 0, 1920, 1080)

        # Inicializar variáveis
        self.active_button = None
        self.active_frases = None
        self.frases = {}
        
        # Dicionário de PDFs organizado (movido para antes do init_ui)
        self.pdf_buttons = {
            "Alteração dos dados do cartão de crédito": "",
            "Card de Propostas Pendentes": "",
            "Chamado - Canal de apoio ao Corretor": "",
            "Chat suporte ao corretor": "",
            "Condições Gerais": "",
            "Critérios de bônus": "",
            "Endosso Online": "",
            "Enquadramento de veículo em tempo de Campanha": "",
            "Manual consulta restrições": "",
            "Manual de extensões de perímetro": "",
            "Manual de PGIT": "",
            "Manual de Regras de Bônus": "",
            "Manual do Corretor": "",
            "Pagamento PGIT": "",
            "Passo a passo Livelo - Corretor": "",
            "Passo a passo Livelo - Matriz": "",
            "Retificação de propostas": "",
            "Retirada de boleto": "",
            "Segunda Chance": "",
            "Tela retificadora do corretor": "",
            "Vistoria Prévia": ""
        }
        
        self.segmentos = {
            "Pagamentos em Andamento": [
                "Boleto",
                "Débito em conta",
                "Cartão de crédito",
                "Endosso - Débito",
                "Endosso - Cartão",
                "CONTRATO - IND"
            ],
            "Segunda Chance": [
                "Boleto - Segunda Chance",
                "Débito em Conta - Segunda Chance",
                "Carto de crédito - Segunda Chance",
                "Endosso Débito - Segunda Chance",
                "Endosso Boleto - Segunda Chance"
            ],
            "Pagamentos e Livelo": [
                "Pagamentos - Pagamentos e Livelo",
                "Livelo - Pagamentos e Livelo"
            ],
            "Classe de Bônus": [
                "Classe de bônus",
                "Canc. Falta de Pagamento",
                "Ren. Sem Sinistro igual ou maior - 335 dias",
                "Ren. Sem Sinistro igual ou menor 335 dias",
                "Leitura de Bônus"
            ],
            "Vistoria": [
                "Posto Fixo",
                "Análise especial"
            ],
            "Rastreador": [
                "Ordem não gerada",
                "Ordem gerada",
                "Ordem cancelada"
            ],
            "Apólice": [
                "Apólice (Geral)",
                "Emissão de boleto",
                "Cartão de crédito"
            ]
        }

        # Inicializar o browser_list aqui
        self.browser_list = QListWidget()
        self.browser_list.setMaximumWidth(400)
        
        # Carregar dados
        self.load_phrases_from_file()
        self.load_pdf_links()
        
        # Inicializar interface
        self.init_ui()
        
        # Carregar estado das marcações
        self.load_marked_items()

        # Configurar atalhos de teclado
        self.shortcut_mark = QShortcut(QKeySequence("Ctrl+M"), self)
        self.shortcut_mark.activated.connect(self.mark_selected_item)

        self.shortcut_unmark = QShortcut(QKeySequence("Ctrl+N"), self)
        self.shortcut_unmark.activated.connect(self.unmark_selected_item)
        
        # Carregar último arquivo TXT ao iniciar
        self.load_txt_file(auto_load=True)

        # Carregar abas salvas
        self.load_tabs()

    def init_ui(self):
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Layout principal com tabs
        self.tab_widget = QTabWidget()
        
        # Primeira aba (Chat)
        primeira_aba = QWidget()
        primeira_layout = QHBoxLayout()
        
        # Container esquerdo
        container_esquerdo = QWidget()
        layout_esquerdo = QVBoxLayout()
        
        # Adiciona o visualizador de site
        site_view = QWebEngineView()
        site_view.setUrl(QUrl("https://www.google.com.br"))
        site_view.setFixedHeight(700)  # Aumenta a altura do visualizador de site
        layout_esquerdo.addWidget(site_view)
        
        # Lista de mensagens com filtro
        filter_layout = self.create_filter_layout()
        layout_esquerdo.addLayout(filter_layout)
        
        self.messages_list = QListWidget()
        self.messages_list.itemClicked.connect(self.copy_to_clipboard)
        # Atualizar esta linha para usar a flag correta do Qt6
        self.messages_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.messages_list.customContextMenuRequested.connect(self.show_context_menu)
        layout_esquerdo.addWidget(self.messages_list)
        
        # Novo frame para botões de alerta
        alert_frame = QFrame()
        # Corrigir as flags do QFrame para PyQt6
        alert_frame.setFrameStyle(QFrame.Shape.Box | QFrame.Shadow.Plain)
        alert_frame.setLineWidth(1)
        alert_buttons_layout = QHBoxLayout(alert_frame)
        alert_buttons_layout.setContentsMargins(10, 5, 10, 5)
        
        # Estilo para os botões
        button_style = """
            QPushButton {
                padding: 5px 15px;
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 3px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
        """
        
        # Criar botões
        novo_email_btn = QPushButton("Novo e-mail")  # Novo botão
        alerta8_btn = QPushButton("Alerta 8")
        alerta9_btn = QPushButton("Alerta 9")
        ppe_btn = QPushButton("PPE")
        sinistro_btn = QPushButton("Sinistro")
        
        # Aplicar estilo aos botões
        for btn in [novo_email_btn, alerta8_btn, alerta9_btn, ppe_btn, sinistro_btn]:  # Incluído novo botão
            btn.setStyleSheet(button_style)
        
        # Adicionar checkbox para controle de imagens, anexos e ausências
        self.remove_images_cb = QCheckBox("Retirar imagens")
        self.remove_attachments_cb = QCheckBox("Remover anexo")
        self.evandro_absent_cb = QCheckBox("Evandro - férias-ausente")
        self.aline_absent_cb = QCheckBox("Aline - férias-ausente")
        
        # Estilo para os checkboxes
        checkbox_style = """
            QCheckBox {
                margin-left: 10px;
                padding: 5px;
            }
        """
        self.remove_images_cb.setStyleSheet(checkbox_style)
        self.remove_attachments_cb.setStyleSheet(checkbox_style)
        self.evandro_absent_cb.setStyleSheet(checkbox_style)
        self.aline_absent_cb.setStyleSheet(checkbox_style)
        
        # Aplicar estilo aos botões
        for btn in [novo_email_btn, alerta8_btn, alerta9_btn, ppe_btn, sinistro_btn]:  # Incluído novo botão
            btn.setStyleSheet(button_style)
        
        # Adicionar botões e checkboxes ao layout
        alert_buttons_layout.addWidget(novo_email_btn)  # Adicionado antes do Alerta 8
        alert_buttons_layout.addWidget(alerta8_btn)
        alert_buttons_layout.addWidget(alerta9_btn)
        alert_buttons_layout.addWidget(ppe_btn)
        alert_buttons_layout.addWidget(sinistro_btn)
        alert_buttons_layout.addWidget(self.remove_images_cb)
        alert_buttons_layout.addWidget(self.remove_attachments_cb)
        alert_buttons_layout.addWidget(self.evandro_absent_cb)
        alert_buttons_layout.addWidget(self.aline_absent_cb)
        alert_buttons_layout.addStretch()
        
        # Conectar sinais dos botões
        novo_email_btn.clicked.connect(self.open_new_email)  # Nova conexão
        alerta8_btn.clicked.connect(lambda: self.open_outlook_email("Alerta 8"))
        alerta9_btn.clicked.connect(lambda: self.open_outlook_email("Alerta 9"))
        ppe_btn.clicked.connect(lambda: self.open_outlook_email("PPE"))
        sinistro_btn.clicked.connect(lambda: self.open_outlook_email("Sinistro"))
        
        # Adicionar frame ao layout
        layout_esquerdo.addWidget(alert_frame)
        
        container_esquerdo.setLayout(layout_esquerdo)
        primeira_layout.addWidget(container_esquerdo, 1)
        
        # Área de botões
        primeira_layout.addWidget(self.create_buttons_area(), 0)
        
        primeira_aba.setLayout(primeira_layout)
        self.tab_widget.addTab(primeira_aba, "Chat")
        
        # Segunda aba (Manuais)
        segunda_aba = QWidget()
        segunda_layout = QHBoxLayout()
        
        # Área do visualizador de PDF (lado esquerdo)
        pdf_viewer_group = QGroupBox("Visualizador")
        pdf_viewer_layout = QVBoxLayout()
        
        # Substituir o QLabel pelo novo PDFViewer
        self.pdf_viewer = PDFViewer()
        pdf_viewer_layout.addWidget(self.pdf_viewer)
        
        pdf_viewer_group.setLayout(pdf_viewer_layout)
        segunda_layout.addWidget(pdf_viewer_group, 4)
        
        # Área de botões (lado direito)
        buttons_group = QGroupBox("Manuais e Documentos")
        buttons_layout = QVBoxLayout()
        
        # Checkbox para modo de edição
        self.edit_mode_cb = QCheckBox("Modo de Edição")
        self.edit_mode_cb.setStyleSheet("""
            QCheckBox {
                padding: 5px;
                font-weight: bold;
            }
        """)
        buttons_layout.addWidget(self.edit_mode_cb)
        
        # Scroll area para os botões
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        # Estilo para os botões
        button_style = """
            QPushButton {
                text-align: left;
                padding: 10px;
                border: 1px solid #ccc;
                border-radius: 5px;
                background-color: #f8f9fa;
            }
            QPushButton:hover {
                background-color: #e9ecef;
            }
        """
        
        # Adicionar botões para cada PDF
        for nome_pdf in self.pdf_buttons.keys():
            btn = QPushButton(nome_pdf)
            btn.setFixedHeight(35)
            btn.setMinimumWidth(50)
            btn.setStyleSheet(button_style)
            btn.clicked.connect(lambda checked, name=nome_pdf: self.handle_pdf_button_click(name))
            scroll_layout.addWidget(btn)
        
        scroll_layout.addStretch()
        scroll_widget.setLayout(scroll_layout)
        
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        buttons_layout.addWidget(scroll)
        
        # Botão de remover
        remove_pdf_button = QPushButton("Remover PDFs")
        remove_pdf_button.clicked.connect(self.show_remove_dialog)
        buttons_layout.addWidget(remove_pdf_button)
        
        buttons_group.setLayout(buttons_layout)
        segunda_layout.addWidget(buttons_group, 1)
        
        segunda_aba.setLayout(segunda_layout)
        self.tab_widget.addTab(segunda_aba, "Manuais")
        
        # Terceira aba (Visualizador TXT)
        terceira_aba = QWidget()
        terceira_layout = QHBoxLayout()
        
        # Área do visualizador de texto (lado esquerdo)
        text_viewer_group = QGroupBox("Visualizador de Texto")
        text_viewer_layout = QVBoxLayout()
        
        # Configuração do QTextEdit com suporte a m��ltipla seleção
        self.text_viewer = QTextEdit()
        self.text_viewer.setReadOnly(True)
        self.text_viewer.setAcceptRichText(False)
        # Corrigir a flag NoWrap para PyQt6
        self.text_viewer.setLineWrapMode(QTextEdit.LineWrapMode.NoWrap)
        self.text_viewer.setUndoRedoEnabled(False)
        self.selected_texts = []
        
        # Configurar o comportamento de seleção múltipla
        self.text_viewer.mousePressEvent = self.custom_mouse_press_event
        self.text_viewer.mouseReleaseEvent = self.custom_mouse_release_event
        
        text_viewer_layout.addWidget(self.text_viewer)
        text_viewer_group.setLayout(text_viewer_layout)
        terceira_layout.addWidget(text_viewer_group, 2)
        
        # Área de botões (lado direito)
        buttons_group = QGroupBox("Controles")
        buttons_layout = QVBoxLayout()
        
        load_button = QPushButton("Carregar Arquivo")
        load_button.clicked.connect(self.load_txt_file)
        
        copy_button = QPushButton("Copiar para Chat")
        copy_button.clicked.connect(self.show_copy_dialog)
        
        buttons_layout.addWidget(load_button)
        buttons_layout.addWidget(copy_button)
        buttons_layout.addStretch()
        
        buttons_group.setLayout(buttons_layout)
        terceira_layout.addWidget(buttons_group, 1)
        
        terceira_aba.setLayout(terceira_layout)
        self.tab_widget.addTab(terceira_aba, "Visualizador TXT")
        
        # Quarta aba (Navegador)
        quarta_aba = QWidget()
        quarta_layout = QVBoxLayout()

        # Criar um QTabWidget para gerenciar múltiplas abas de sites
        self.browser_tabs = QTabWidget()
        self.browser_tabs.setTabsClosable(True)
        self.browser_tabs.setMovable(True)
        self.browser_tabs.tabCloseRequested.connect(self.close_browser_tab)
        self.browser_tabs.tabBarClicked.connect(self.handle_tab_click)

        # Adicionar o QTabWidget ao layout da aba "Navegador"
        quarta_layout.addWidget(self.browser_tabs)

        quarta_aba.setLayout(quarta_layout)
        self.tab_widget.addTab(quarta_aba, "Navegador")
        
        # Quinta aba (Browser)
        quinta_aba = QWidget()
        quinta_layout = QHBoxLayout()  # Layout horizontal

        # Container esquerdo (lista de sites)
        left_container = QWidget()
        left_container.setMaximumWidth(250)  # Limita largura do container esquerdo
        left_layout = QVBoxLayout()

        # Botões de controle
        button_style = """
            QPushButton {
                padding: 8px;
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                margin: 2px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """

        # Criar botões
        add_btn = QPushButton("Adicionar")
        remove_btn = QPushButton("Remover")
        edit_btn = QPushButton("Editar")
        export_btn = QPushButton("Exportar")
        import_btn = QPushButton("Importar")

        # Conectar os botões às funções
        add_btn.clicked.connect(self.add_site_to_list)  # Este já funciona
        remove_btn.clicked.connect(self.show_remove_site_dialog)
        edit_btn.clicked.connect(self.show_edit_site_dialog)
        export_btn.clicked.connect(lambda: QFileDialog.getSaveFileName(
            self,
            "Exportar Sites",
            "",
            "CSV Files (*.csv)"
        ))
        import_btn.clicked.connect(lambda: QFileDialog.getOpenFileName(
            self,
            "Importar Sites",
            "",
            "CSV Files (*.csv)"
        ))

        # Aplicar estilo aos botões
        for btn in [add_btn, remove_btn, edit_btn, export_btn, import_btn]:
            btn.setStyleSheet(button_style)
            btn.setFixedHeight(40)
            left_layout.addWidget(btn)

        # Lista de sites
        self.sites_list = QListWidget()
        self.sites_list.itemClicked.connect(self.on_site_selected)
        left_layout.addWidget(self.sites_list)

        left_container.setLayout(left_layout)

        # Container direito (visualizador web)
        right_container = QWidget()
        right_layout = QVBoxLayout()

        # Criar QTabWidget para as abas de sites
        self.sites_tabs = QTabWidget()
        self.sites_tabs.setTabsClosable(True)
        self.sites_tabs.setMovable(True)
        self.sites_tabs.tabCloseRequested.connect(self.close_site_tab)

        # Adicionar QTabWidget ao layout direito
        right_layout.addWidget(self.sites_tabs)
        right_container.setLayout(right_layout)

        # Adicionar os containers ao layout principal
        quinta_layout.addWidget(left_container)
        quinta_layout.addWidget(right_container, stretch=1)  # stretch=1 para ocupar mais espaço

        quinta_aba.setLayout(quinta_layout)
        self.tab_widget.addTab(quinta_aba, "Browser")

        # Carregar sites salvos
        self.load_sites()
        
        # Layout principal
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.tab_widget)
        central_widget.setLayout(main_layout)
        
        # Barra de status
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

    def set_active_button(self, button, frases):
        """Define o botão ativo e altera seu estilo."""
        # Verifica se já existe um botão ativo e restaura o estilo padrão
        if self.active_button is not None:
            try:
                self.active_button.setStyleSheet("")
            except RuntimeError:
                # Se o botão ativo anterior foi deletado, redefina a referência
                self.active_button = None

        # Define o novo botão ativo
        self.active_button = button
        self.active_button.setStyleSheet("background-color: lightblue;")

        # Atualizar frases ativas
        self.active_frases = frases

        # Atualizar lista de mensagens
        self.messages_list.clear()
        for frase in sorted(self.active_frases):
            self.messages_list.addItem(QListWidgetItem(frase))

        # Reaplicar marcações após atualizar a lista
        self.load_marked_items()

    def copy_to_clipboard(self, item):
        """Copia a frase para a área de transferência"""
        clipboard = QApplication.clipboard()
        clipboard.setText(item.text())

    def load_phrases_from_file(self):
        try:
            with open('frases.json', 'r', encoding='utf-8') as file:
                data = json.load(file)
                self.frases = data.get('frases', {})
                loaded_segmentos = data.get('segmentos', {})
                if loaded_segmentos:
                    self.segmentos.update(loaded_segmentos)
        except FileNotFoundError:
            print("Arquivo de frases não encontrado. Criando novo arquivo.")
            self.save_phrases_to_file()
        except Exception as e:
            print(f"Erro ao carregar frases: {str(e)}")

    def save_phrases_to_file(self):
        try:
            data = {
                'frases': self.frases,
                'segmentos': self.segmentos
            }
            with open('frases.json', 'w', encoding='utf-8') as file:
                json.dump(data, file, ensure_ascii=False, indent=4)
            return True
        except Exception as e:
            print(f"Erro ao salvar frases: {str(e)}")
            return False

    def create_buttons_area(self):
        """Cria a área de botões"""
        container = QWidget()
        main_layout = QVBoxLayout(container)
        
        # Ajustando margens do layout principal
        main_layout.setContentsMargins(25, 15, 25, 15)
        main_layout.setSpacing(8)
        
        # Botões de controle em um widget fixo
        control_widget = QWidget()
        control_layout = QHBoxLayout(control_widget)
        control_widget.setFixedHeight(50)  # Altura fixa para os botões
        
        add_button = QPushButton("+")
        remove_button = QPushButton("-")
        edit_button = QPushButton("Edição")
        add_button.setFixedSize(30, 30)
        remove_button.setFixedSize(30, 30)
        edit_button.setFixedSize(60, 30)
        add_button.clicked.connect(self.add_button_clicked)
        remove_button.clicked.connect(self.remove_button_clicked)
        edit_button.clicked.connect(self.edit_button_clicked)
        control_layout.addWidget(add_button)
        control_layout.addWidget(remove_button)
        control_layout.addWidget(edit_button)
        control_layout.addStretch()
        
        # Área de scroll para os segmentos e botões
        scroll = QScrollArea()
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        
        # Criar botões por segmento
        for segmento, botoes in self.segmentos.items():
            label = QLabel(segmento)
            label.setStyleSheet("font-weight: bold; margin-top: 5px;")
            scroll_layout.addWidget(label)
            
            section_layout = QVBoxLayout()
            section_layout.setSpacing(8)
            section_layout.setContentsMargins(5, 0, 5, 0)
            
            for botao in botoes:
                if botao not in self.frases:
                    self.frases[botao] = []
                
                btn = QPushButton(botao)
                btn.setFixedSize(340, 50)
                btn.clicked.connect(lambda checked, b=btn, f=self.frases[botao]: 
                                  self.set_active_button(b, f))
                section_layout.addWidget(btn)
            
            scroll_layout.addLayout(section_layout)
            scroll_layout.addSpacing(5)
        
        scroll_content.setLayout(scroll_layout)
        scroll.setWidget(scroll_content)
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet("""
            QScrollArea {
                border: 1px solid #999;
                border-radius: 0px;
                background-color: white;
                margin: 5px;
            }
        """)
        
        # Adicionar os widgets ao layout principal
        main_layout.addWidget(control_widget)  # Botões de controle fixos no topo
        main_layout.addWidget(scroll)          # Área de scroll abaixo
        
        return container

    def add_button_clicked(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Adicionar Novo Botão")
        dialog.setMinimumWidth(600)
        dialog.setMinimumHeight(400)
        
        # Layout principal
        final_layout = QVBoxLayout(dialog)
        
        # Container principal
        main_layout = QHBoxLayout()
        
        # Layout esquerdo (sempre visível)
        left_layout = QVBoxLayout()
        
        # Container para "Novo Segmento"
        new_segment_container = QVBoxLayout()
        checkbox_layout = QHBoxLayout()
        new_segment_checkbox = QCheckBox()
        new_segment_label = QLabel("Novo Segmento")
        new_segment_label.setStyleSheet("font-weight: bold;")
        checkbox_layout.addWidget(new_segment_checkbox)
        checkbox_layout.addWidget(new_segment_label)
        checkbox_layout.addStretch()
        new_segment_container.addLayout(checkbox_layout)
        
        segment_input = QLineEdit()
        segment_input.setPlaceholderText("Digite o nome do novo segmento")
        segment_input.hide()
        new_segment_container.addWidget(segment_input)
        
        button_label = QLabel("Nome do Botão:")
        button_input = QLineEdit()
        button_input.setPlaceholderText("Digite o nome do botão")
        new_segment_container.addWidget(button_label)
        new_segment_container.addWidget(button_input)
        
        add_item_button = QPushButton("Adicionar")
        add_item_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                padding: 5px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        new_segment_container.addWidget(add_item_button)
        
        left_layout.addLayout(new_segment_container)
        left_layout.addStretch()
        
        # Separador vertical
        separator = QFrame()
        separator.setFrameShape(QFrame.Shape.VLine)
        separator.setFrameShadow(QFrame.Shadow.Sunken)
        
        # Área direita com scroll
        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        right_content = QWidget()
        right_layout = QVBoxLayout(right_content)
        
        # Lista de segmentos existentes
        for segmento, botoes in self.segmentos.items():
            segment_container = QHBoxLayout()
            segment_checkbox = QCheckBox()
            segment_container.addWidget(segment_checkbox)
            segment_label = QLabel(f"{segmento}:")
            segment_label.setStyleSheet("font-weight: bold;")
            segment_container.addWidget(segment_label)
            segment_container.addStretch()
            right_layout.addLayout(segment_container)
            
            for botao in botoes:
                button_container = QHBoxLayout()
                button_container.addSpacing(20)
                button_label = QLabel(botao)
                button_container.addWidget(button_label)
                button_container.addStretch()
                right_layout.addLayout(button_container)
        
        right_scroll.setWidget(right_content)
        
        # Adiciona os layouts ao main_layout
        main_layout.addLayout(left_layout, 1)
        main_layout.addWidget(separator)
        main_layout.addWidget(right_scroll, 2)
        
        # Botões OK/Cancelar
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | 
            QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        
        # Adiciona widgets ao layout final
        final_layout.addLayout(main_layout)
        final_layout.addWidget(button_box)
        
        # Função para adicionar item com atualização imediata
        def add_item():
            button_name = button_input.text().strip()
            if button_name:
                if new_segment_checkbox.isChecked():
                    segment_name = segment_input.text().strip()
                    if segment_name:
                        if segment_name not in self.segmentos:
                            self.segmentos[segment_name] = []
                            # Adicionar novo segmento à visualização
                            segment_container = QHBoxLayout()
                            segment_checkbox = QCheckBox()
                            segment_container.addWidget(segment_checkbox)
                            segment_label = QLabel(f"{segment_name}:")
                            segment_label.setStyleSheet("font-weight: bold;")
                            segment_container.addWidget(segment_label)
                            segment_container.addStretch()
                            right_layout.addLayout(segment_container)
                        
                        if button_name not in self.segmentos[segment_name]:
                            self.segmentos[segment_name].append(button_name)
                            self.frases[button_name] = []
                            
                            # Adicionar novo botão à visualização
                            button_container = QHBoxLayout()
                            button_container.addSpacing(20)
                            button_label = QLabel(button_name)
                            button_container.addWidget(button_label)
                            button_container.addStretch()
                            
                            # Encontrar a posição correta para inserção
                            insert_position = right_layout.count() - 1
                            while insert_position >= 0:
                                next_item = right_layout.itemAt(insert_position)
                                if isinstance(next_item, QHBoxLayout):
                                    next_widget = next_item.itemAt(0).widget()
                                    if isinstance(next_widget, QCheckBox):
                                        break
                                insert_position -= 1
                            
                            # Inserir o novo botão
                            right_layout.insertLayout(insert_position + 1, button_container)
                            
                            button_input.clear()
                            segment_input.clear()
                            QMessageBox.information(dialog, "Sucesso", "Item adicionado com sucesso!")
                        else:
                            QMessageBox.warning(dialog, "Aviso", "Este botão já existe neste segmento!")
                    else:
                        QMessageBox.warning(dialog, "Aviso", "Digite o nome do segmento!")
                else:
                    # Coletar todos os segmentos selecionados
                    selected_segments = []
                    for i in range(right_layout.count()):
                        item = right_layout.itemAt(i)
                        if isinstance(item, QHBoxLayout):
                            checkbox = item.itemAt(0).widget()
                            label = item.itemAt(1).widget()
                            if isinstance(checkbox, QCheckBox) and checkbox.isChecked():
                                segment_name = label.text().replace(':', '')
                                selected_segments.append((segment_name, i))
                    
                    if selected_segments:
                        # Adicionar o botão em todos os segmentos selecionados
                        offset = 0  # Para ajustar as posições após cada inserção
                        for segment_name, index in selected_segments:
                            if button_name not in self.segmentos[segment_name]:
                                # Adicionar aos dados
                                self.segmentos[segment_name].append(button_name)
                                if button_name not in self.frases:
                                    self.frases[button_name] = []
                                
                                # Criar o novo botão
                                button_container = QHBoxLayout()
                                button_container.addSpacing(20)
                                button_label = QLabel(button_name)
                                button_container.addWidget(button_label)
                                button_container.addStretch()
                                
                                # Encontrar a posição correta para inserção
                                insert_position = index + offset + 1
                                while insert_position < right_layout.count():
                                    next_item = right_layout.itemAt(insert_position)
                                    if isinstance(next_item, QHBoxLayout):
                                        next_widget = next_item.itemAt(0).widget()
                                        if isinstance(next_widget, QCheckBox):
                                            break
                                    insert_position += 1
                                
                                # Inserir o novo botão
                                right_layout.insertLayout(insert_position, button_container)
                                offset += 1  # Incrementar o offset para a próxima inserção
                        
                        button_input.clear()
                        # Forçar atualização visual
                        right_content.updateGeometry()
                        right_scroll.updateGeometry()
                        QMessageBox.information(dialog, "Sucesso", "Item adicionado com sucesso!")
                    else:
                        QMessageBox.warning(dialog, "Aviso", "Selecione pelo menos um segmento!")
            else:
                QMessageBox.warning(dialog, "Aviso", "Digite o nome do botão!")
            
            # Forçar atualização visual
            right_content.updateGeometry()
            right_scroll.updateGeometry()
        
        # Conectar botão à função
        add_item_button.clicked.connect(add_item)
        
        # Funço para mostrar/ocultar campo de segmento
        def toggle_segment_input():
            segment_input.setVisible(new_segment_checkbox.isChecked())
            dialog.adjustSize()
        
        new_segment_checkbox.stateChanged.connect(toggle_segment_input)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.save_phrases_to_file()
            self.init_ui()

    def remove_button_clicked(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Remover")
        dialog.setMinimumWidth(400)
        dialog.setMinimumHeight(300)
        
        # Layout principal
        final_layout = QVBoxLayout(dialog)
        
        # Scroll Area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        # Widget para conter o conteúdo scrollável
        scroll_content = QWidget()
        content_layout = QVBoxLayout(scroll_content)

        # Dicionário para armazenar os widgets
        widgets_map = {}

        # Lista de segmentos e botões com checkboxes
        for segmento, botoes in self.segmentos.items():
            # Container para o segmento
            segment_widget = QWidget()
            segment_layout = QVBoxLayout(segment_widget)
            
            # Container horizontal para checkbox e label do segmento
            segment_container = QHBoxLayout()
            
            # Checkbox para o segmento
            segment_checkbox = QCheckBox()
            segment_container.addWidget(segment_checkbox)
            
            # Label do segmento
            segment_label = QLabel(f"{segmento}:")
            segment_label.setStyleSheet("font-weight: bold;")
            segment_container.addWidget(segment_label)
            segment_container.addStretch()
            
            segment_layout.addLayout(segment_container)
            
            # Armazenar widgets do segmento
            widgets_map[segmento] = {
                'checkbox': segment_checkbox,
                'container': segment_widget,
                'botoes': {}
            }
            
            # Checkboxes para os botões (com indentação)
            for botao in botoes:
                button_widget = QWidget()
                button_container = QHBoxLayout(button_widget)
                button_container.addSpacing(20)  # Indentação
                
                # Checkbox do botão
                button_checkbox = QCheckBox()
                button_container.addWidget(button_checkbox)
                
                # Label do botão
                button_label = QLabel(botao)
                button_container.addWidget(button_label)
                button_container.addStretch()
                
                segment_layout.addWidget(button_widget)
                
                # Armazenar widgets do botão
                widgets_map[segmento]['botoes'][botao] = {
                    'checkbox': button_checkbox,
                    'container': button_widget
                }
            
            content_layout.addWidget(segment_widget)

        def remove_selected_items():
            nonlocal widgets_map
            items_removed = False
            has_selected = False
            
            # Verificar se há itens selecionados
            for segmento, widgets in widgets_map.items():
                if widgets['checkbox'].isChecked():
                    has_selected = True
                    break
                for botao, button_widgets in widgets['botoes'].items():
                    if button_widgets['checkbox'].isChecked():
                        has_selected = True
                        break
                if has_selected:
                    break
            
            if not has_selected:
                QMessageBox.warning(dialog, "Aviso", "Selecione pelo menos um item para remover!")
                return
            
            # Confirmação antes de remover
            confirm_msg = QMessageBox(dialog)
            confirm_msg.setIcon(QMessageBox.Icon.Question)
            confirm_msg.setWindowTitle("Confirmar Remoção")
            confirm_msg.setText("Deseja remover os itens selecionados?")
            confirm_msg.setStandardButtons(
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if confirm_msg.exec() == QMessageBox.StandardButton.Yes:
                try:
                    # Remover segmentos marcados
                    for segmento, widgets in dict(widgets_map).items():
                        segment_checkbox = widgets['checkbox']
                        if segment_checkbox.isChecked():
                            # Remover frases dos botões do segmento
                            for botao in self.segmentos[segmento]:
                                if botao in self.frases:
                                    del self.frases[botao]
                            
                            # Remover segmento
                            del self.segmentos[segmento]
                            
                            # Remover widgets do segmento
                            for widget in widgets['container'].children():
                                widget.deleteLater()
                            widgets['container'].deleteLater()
                            del widgets_map[segmento]
                            items_removed = True
                            continue
                        
                        # Remover botões individuais marcados
                        buttons_to_remove = []
                        for botao, button_widgets in widgets['botoes'].items():
                            if button_widgets['checkbox'].isChecked():
                                buttons_to_remove.append(botao)
                                
                        for botao in buttons_to_remove:
                            # Remover botão do segmento
                            self.segmentos[segmento].remove(botao)
                            # Remover frases do botão
                            if botao in self.frases:
                                del self.frases[botao]
                            # Remover widgets do botão
                            button_container = widgets['botoes'][botao]['container']
                            for widget in button_container.children():
                                widget.deleteLater()
                            button_container.deleteLater()
                            del widgets['botoes'][botao]
                            items_removed = True
                        
                        # Se o segmento ficou vazio, removê-lo
                        if segmento in self.segmentos and not self.segmentos[segmento]:
                            del self.segmentos[segmento]
                            for widget in widgets['container'].children():
                                widget.deleteLater()
                            widgets['container'].deleteLater()
                            del widgets_map[segmento]
                    
                    if items_removed:
                        self.save_phrases_to_file()
                        QMessageBox.information(dialog, "Sucesso", "Itens removidos com sucesso!")
                        scroll_content.adjustSize()
                        
                except Exception as e:
                    QMessageBox.warning(dialog, "Erro", f"Erro ao remover itens: {str(e)}")

        # Botão Remover Selecionados
        remove_button = QPushButton("Remover Selecionados")
        remove_button.clicked.connect(remove_selected_items)
        remove_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                padding: 5px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        
        # Configura a scroll area
        scroll_area.setWidget(scroll_content)

        # Botões de confirmação
        button_box = QDialogButtonBox()
        ok_button = button_box.addButton(QDialogButtonBox.StandardButton.Ok)
        cancel_button = button_box.addButton(QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)

        # Adiciona widgets ao layout final
        final_layout.addWidget(scroll_area)
        final_layout.addWidget(remove_button)
        final_layout.addWidget(button_box)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.init_ui()

    def edit_button_clicked(self):
        """Função para editar botões"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Editar Botões")
        dialog.setMinimumWidth(400)
        layout = QVBoxLayout()

        # Scroll area para os checkboxes e labels editáveis
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout()

        # Dicionário para armazenar os widgets de edição
        edit_widgets = {}

        # Criar widgets de edição para cada segmento e botão
        for segmento, botoes in self.segmentos.items():
            segment_label = QLabel(segmento)
            segment_label.setStyleSheet("font-weight: bold;")
            scroll_layout.addWidget(segment_label)

            for botao in botoes:
                button_layout = QHBoxLayout()
                
                checkbox = QCheckBox()
                button_layout.addWidget(checkbox)
                
                label = QLineEdit(botao)
                label.setReadOnly(True)
                label.setStyleSheet("background-color: #f0f0f0;")
                button_layout.addWidget(label)
                
                # Conectar checkbox com a função de toggle da label
                checkbox.stateChanged.connect(
                    lambda state, lbl=label: lbl.setReadOnly(not state))  # Fechando o parêntese aqui
                
                edit_widgets[botao] = {
                    'checkbox': checkbox,
                    'label': label,
                    'segmento': segmento
                }
                
                scroll_layout.addLayout(button_layout)

        scroll_widget.setLayout(scroll_layout)
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        layout.addWidget(scroll)

        # Botões OK/Cancelar
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | 
            QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        dialog.setLayout(layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Processar alterações
            for botao, widgets in edit_widgets.items():
                if widgets['checkbox'].isChecked():
                    novo_nome = widgets['label'].text().strip()
                    if novo_nome and novo_nome != botao:
                        segmento = widgets['segmento']
                        # Atualizar nome no segmento
                        idx = self.segmentos[segmento].index(botao)
                        self.segmentos[segmento][idx] = novo_nome
                        # Atualizar frases
                        if botao in self.frases:
                            self.frases[novo_nome] = self.frases.pop(botao)
            
            self.save_phrases_to_file()
            self.init_ui()

    def update_buttons_area(self):
        """Atualiza a área de botões após modificações"""
        primeira_aba = self.tab_widget.widget(0)
        if primeira_aba and primeira_aba.layout():
            layout = primeira_aba.layout()
            
            # Remover área de botões antiga
            if layout.count() > 1:
                old_buttons = layout.takeAt(1)
                if old_buttons and old_buttons.widget():
                    old_buttons.widget().deleteLater()
            
            # Adicionar nova área de botões
            layout.addWidget(self.create_buttons_area(), 0)

    def add_pdf(self, pdf_name=None):
        """Adiciona ou atualiza um PDF"""
        if pdf_name is None:
            pdf_name, ok = QInputDialog.getItem(
                self,
                "Selecionar Manual",
                "Escolha o manual para adicionar:",
                [k for k, v in self.pdf_buttons.items() if not v],
                0,
                False
            )
            if not ok:
                return

        file_name, _ = QFileDialog.getOpenFileName(
            self,
            f"Selecionar PDF para {pdf_name}",
            "",
            "PDF Files (*.pdf)"
        )
        
        if file_name:
            try:
                # Criar diretório pdfs se não existir
                if not os.path.exists('pdfs'):
                    os.makedirs('pdfs')
                
                # Copiar arquivo para o diretório pdfs
                new_name = f"{pdf_name}.pdf"
                new_path = os.path.join('pdfs', new_name)
                shutil.copy2(file_name, new_path)
                
                # Atualizar dicionário com caminho absoluto
                self.pdf_buttons[pdf_name] = os.path.abspath(new_path)
                self.save_pdf_links()
                
                # Tentar abrir o PDF imediatamente
                self.pdf_viewer.load_pdf(self.pdf_buttons[pdf_name])
                
                self.status_bar.showMessage(f"PDF '{pdf_name}' adicionado com sucesso!", 2000)
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao adicionar PDF: {str(e)}")

    def remove_pdf(self):
        """Remove o vínculo com um PDF"""
        pdf_name, ok = QInputDialog.getItem(
            self,
            "Remover PDF",
            "Escolha o manual para remover:",
            [k for k, v in self.pdf_buttons.items() if v],  # Mostrar apenas os com arquivo
            0,
            False
        )
        if ok and pdf_name:
            try:
                pdf_path = self.pdf_buttons[pdf_name]
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                
                self.pdf_buttons[pdf_name] = ""
                self.save_pdf_links()
                
                self.status_bar.showMessage(f"PDF '{pdf_name}' removido com sucesso!", 2000)
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao remover PDF: {str(e)}")

    def save_pdf_links(self):
        """Salva os links dos PDFs em um arquivo"""
        try:
            with open('pdf_links.json', 'w', encoding='utf-8') as f:
                json.dump(self.pdf_buttons, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Erro ao salvar links dos PDFs: {str(e)}")

    def load_pdf_links(self):
        """Carrega os links dos PDFs do arquivo"""
        try:
            if os.path.exists('pdf_links.json'):
                with open('pdf_links.json', 'r', encoding='utf-8') as f:
                    saved_links = json.load(f)
                    self.pdf_buttons.update(saved_links)
        except Exception as e:
            print(f"Erro ao carregar links dos PDFs: {str(e)}")

    def open_pdf(self, pdf_name):
        """Abre o PDF selecionado"""
        pdf_path = self.pdf_buttons.get(pdf_name)
        print(f"Tentando abrir PDF: {pdf_name}")
        print(f"Caminho do arquivo: {pdf_path}")
        
        if not pdf_path:
            QMessageBox.warning(
                self,
                "PDF não encontrado",
                f"O PDF '{pdf_name}' ainda não foi vinculado. Por favor, adicione o arquivo."
            )
            if self.edit_mode_cb.isChecked():
                self.add_pdf(pdf_name)
        else:
            try:
                if os.path.exists(pdf_path):
                    print(f"Arquivo existe em: {pdf_path}")
                    # Carregar PDF no visualizador
                    self.pdf_viewer.load_pdf(pdf_path)
                    self.status_bar.showMessage(f"PDF carregado: {pdf_name}", 2000)
                else:
                    print(f"Arquivo não encontrado em: {pdf_path}")
                    QMessageBox.warning(self, "Erro", "Arquivo PDF não encontrado!")
            except Exception as e:
                print(f"Erro ao abrir PDF: {str(e)}")
                QMessageBox.warning(self, "Erro", f"Erro ao abrir PDF: {str(e)}")

    def update_button_combo(self, segment):
        """Atualiza o combo de botões baseado no segmento selecionado"""
        self.button_combo.clear()
        if segment in self.segmentos:
            self.button_combo.addItems(self.segmentos[segment])
        self.update_edit_list()

    def update_edit_list(self):
        """Atualiza a lista de frases do botão selecionado"""
        self.edit_list.clear()
        button = self.button_combo.currentText()
        if button in self.frases:
            self.edit_list.addItems(self.frases[button])

    def add_phrase(self):
        """Adiciona uma nova frase ao botão selecionado"""
        button = self.button_combo.currentText()
        if button:
            text, ok = QInputDialog.getText(self, 'Adicionar Frase', 'Digite a nova frase:')
            if ok and text:
                if button not in self.frases:
                    self.frases[button] = []
                self.frases[button].append(text)
                self.save_phrases_to_file()
                self.update_edit_list()
                self.status_bar.showMessage("Frase adicionada com sucesso!", 2000)

    def edit_phrase(self):
        """Edita a frase selecionada"""
        current_item = self.edit_list.currentItem()
        if current_item:
            button = self.button_combo.currentText()
            old_text = current_item.text()
            text, ok = QInputDialog.getText(self, 'Editar Frase', 'Edite a frase:', text=old_text)
            if ok and text:
                idx = self.frases[button].index(old_text)
                self.frases[button][idx] = text
                self.save_phrases_to_file()
                self.update_edit_list()
                self.status_bar.showMessage("Frase editada com sucesso!", 2000)

    def remove_phrase(self):
        """Remove a frase selecionada"""
        current_item = self.edit_list.currentItem()
        if current_item:
            button = self.button_combo.currentText()
            text = current_item.text()
            reply = QMessageBox.question(self, 'Confirmar Remoção', 
                                       'Tem certeza que deseja remover esta frase?',
                                       QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.frases[button].remove(text)
                self.save_phrases_to_file()
                self.update_edit_list()
                self.status_bar.showMessage("Frase removida com sucesso!", 2000)

    def open_text_file(self):
        """Abre um arquivo de texto"""
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Abrir Arquivo", "", "Arquivos de Texto (*.txt);;Todos os Arquivos (*)"
        )
        if file_name:
            try:
                with open(file_name, 'r', encoding='utf-8') as file:
                    self.edit_text.setPlainText(file.read())
                self.status_bar.showMessage(f"Arquivo aberto: {file_name}", 2000)
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao abrir arquivo: {str(e)}")

    def save_text_file(self):
        """Salva o arquivo de texto"""
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Salvar Arquivo", "", "Arquivos de Texto (*.txt);;Todos os Arquivos (*)"
        )
        if file_name:
            try:
                with open(file_name, 'w', encoding='utf-8') as file:
                    file.write(self.edit_text.toPlainText())
                self.status_bar.showMessage(f"Arquivo salvo: {file_name}", 2000)
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao salvar arquivo: {str(e)}")

    def handle_pdf_button_click(self, pdf_name):
        """Gerencia o clique nos botões de PDF"""
        if self.edit_mode_cb.isChecked():
            self.add_pdf(pdf_name)
        else:
            self.open_pdf(pdf_name)

    def show_remove_dialog(self):
        """Mostra diálogo para remover PDFs"""
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("Remover PDFs")
            dialog.setMinimumWidth(400)
            layout = QVBoxLayout()

            # Lista de PDFs com checkboxes
            scroll = QScrollArea()
            scroll_content = QWidget()
            scroll_layout = QVBoxLayout(scroll_content)
            
            # Lista para armazenar todos os checkboxes
            all_checkboxes = []
            
            # Criar checkbox para cada botão, indicando se tem PDF vinculado
            for pdf_name, pdf_path in self.pdf_buttons.items():
                cb = QCheckBox(pdf_name)
                if pdf_path:
                    cb.setText(f"{pdf_name} (PDF vinculado)")
                    cb.setStyleSheet("QCheckBox { color: green; }")
                else:
                    cb.setText(f"{pdf_name} (Sem PDF)")
                    cb.setStyleSheet("QCheckBox { color: gray; }")
                    cb.setEnabled(False)  # Desabilita checkbox para botões sem PDF
                all_checkboxes.append(cb)
                scroll_layout.addWidget(cb)
            
            scroll_content.setLayout(scroll_layout)
            scroll.setWidget(scroll_content)
            scroll.setWidgetResizable(True)
            scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
            
            # Botões de seleção
            select_buttons = QHBoxLayout()
            select_all = QPushButton("Selecionar Todos")
            deselect_all = QPushButton("Desselecionar Todos")
            
            def select_all_boxes():
                for checkbox in all_checkboxes:
                    if checkbox.isEnabled():  # Seleciona apenas os que têm PDF
                        checkbox.setChecked(True)
            
            def deselect_all_boxes():
                for checkbox in all_checkboxes:
                    checkbox.setChecked(False)
            
            select_all.clicked.connect(select_all_boxes)
            deselect_all.clicked.connect(deselect_all_boxes)
            
            select_buttons.addWidget(select_all)
            select_buttons.addWidget(deselect_all)
            
            # Botões OK/Cancelar
            buttons = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Ok | 
                QDialogButtonBox.StandardButton.Cancel
            )
            buttons.accepted.connect(dialog.accept)
            buttons.rejected.connect(dialog.reject)
            
            layout.addWidget(scroll)
            layout.addLayout(select_buttons)
            layout.addWidget(buttons)
            dialog.setLayout(layout)
            
            if dialog.exec() == QDialog.DialogCode.Accepted:
                # Coletar PDFs selecionados
                selected_pdfs = [
                    pdf_name for pdf_name, cb in zip(self.pdf_buttons.keys(), all_checkboxes)
                    if cb.isChecked() and self.pdf_buttons[pdf_name]
                ]
                
                if selected_pdfs:
                    # Confirmar remoção
                    confirm = QMessageBox.question(
                        self,
                        "Confirmar Remoção",
                        f"Deseja remover {len(selected_pdfs)} PDF(s)?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    
                    if confirm == QMessageBox.StandardButton.Yes:
                        # Remover PDFs selecionados
                        for pdf_name in selected_pdfs:
                            self.pdf_buttons[pdf_name] = ""
                        
                        QMessageBox.information(
                            self,
                            "Sucesso",
                            f"{len(selected_pdfs)} PDF(s) removido(s) com sucesso!"
                        )
                        
                        # Atualizar interface
                        self.save_pdf_links()
                        
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao remover PDFs: {str(e)}")

    def create_filter_layout(self):
        filter_layout = QHBoxLayout()
        
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Digite para filtrar as mensagens...")
        self.filter_input.textChanged.connect(self.filter_messages)
        
        add_button = QPushButton("+")
        add_button.setFixedSize(30, 30)
        add_button.clicked.connect(self.add_frase)
        
        remove_button = QPushButton("-")
        remove_button.setFixedSize(30, 30)
        remove_button.clicked.connect(self.remove_frase)
        
        view_button = QPushButton("Ver Planilha")
        view_button.setFixedSize(100, 30)
        view_button.clicked.connect(self.view_planilha)
        
        # Novo botão Atualizar
        update_button = QPushButton("Atualizar")
        update_button.setFixedSize(80, 30)
        update_button.clicked.connect(self.update_excel)
        
        export_button = QPushButton("Exportar para Excel")
        export_button.setFixedSize(150, 30)
        export_button.clicked.connect(self.export_to_excel)
        
        open_folder_button = QPushButton("Abrir")
        open_folder_button.setFixedSize(50, 30)
        open_folder_button.clicked.connect(self.open_export_folder)
        
        filter_layout.addWidget(self.filter_input)
        filter_layout.addWidget(add_button)
        filter_layout.addWidget(remove_button)
        filter_layout.addWidget(view_button)
        filter_layout.addWidget(update_button)  # Adicionando o novo botão
        filter_layout.addWidget(export_button)
        filter_layout.addWidget(open_folder_button)
        
        return filter_layout

    def filter_messages(self, text):
        for index in range(self.messages_list.count()):
            item = self.messages_list.item(index)
            item.setHidden(text.lower() not in item.text().lower())

    def view_planilha(self):
        """Abre uma janela de diálogo para visualizar as frases em formato de tabela"""
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("Planilha de Dados")
            dialog_layout = QVBoxLayout()

            if self.active_frases:
                # Criar tabela
                table = QTableWidget()
                table.setRowCount(len(self.active_frases))
                table.setColumnCount(1)
                table.setHorizontalHeaderLabels(["Frases"])
                table.horizontalHeader().setStretchLastSection(True)

                # Preencher tabela com as frases
                for row, frase in enumerate(sorted(self.active_frases)):
                    table.setItem(row, 0, QTableWidgetItem(frase))

                # Configurar a tabela para permitir edição
                table.cellChanged.connect(lambda row, col: self.update_phrase(row, col, table))
                
                dialog_layout.addWidget(table)
                
                # Criar botões usando QPushButton
                button_layout = QHBoxLayout()
                ok_button = QPushButton("OK")
                ok_button.clicked.connect(dialog.accept)
                button_layout.addStretch()
                button_layout.addWidget(ok_button)
                
                dialog_layout.addLayout(button_layout)
                dialog.setLayout(dialog_layout)
                dialog.resize(600, 400)
                
                dialog.exec()
            else:
                QMessageBox.information(self, "Aviso", "Nenhuma frase disponível para visualização.")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Ocorreu um erro ao abrir a planilha: {str(e)}")

    def update_phrase(self, row, col, table):
        """Atualiza a frase quando editada na tabela"""
        if self.active_frases and row < len(self.active_frases):
            item = table.item(row, col)
            if item:
                new_text = item.text()
                self.active_frases[row] = new_text
                self.active_frases = sorted(self.active_frases)
                self.frases[self.active_button.text()] = self.active_frases
                self.mostrar_frases()
                self.save_phrases_to_file()

    def export_to_excel(self):
        """Exporta as frases para um arquivo Excel"""
        try:
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Exportar para Excel",
                "",
                "Arquivos Excel (*.xlsx)"
            )
            
            if not file_name:
                return
            
            if not file_name.endswith('.xlsx'):
                file_name += '.xlsx'

            workbook = Workbook()
            
            for segmento, botoes in self.segmentos.items():
                for botao in botoes:
                    if botao in self.frases and self.frases[botao]:
                        sheet_name = botao[:31].upper()  # Nome em maiúsculas
                        if sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                        else:
                            sheet = workbook.create_sheet(title=sheet_name)
                        
                        # Título em maiúsculas
                        header_cell = sheet.cell(row=1, column=1, value=botao.upper())
                        header_cell.font = Font(name='Calibri', size=8, bold=True)
                        header_cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Calcular largura inicial
                        max_length = len(botao.upper())
                        for frase in self.frases[botao]:
                            max_length = max(max_length, len(str(frase)))
                        sheet.column_dimensions['A'].width = (max_length + 2) * 1.2
                        
                        for idx, frase in enumerate(sorted(self.frases[botao]), start=2):
                            cell = sheet.cell(row=idx, column=1, value=str(frase))
                            cell.font = Font(name='Calibri', size=8)
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        for row in sheet.rows:
                            sheet.row_dimensions[row[0].row].height = 15

            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
                
            # Salvar arquivo temporariamente
            workbook.save(file_name)
            
            # Aplicar AutoFit usando win32com
            import win32com.client
            xl = win32com.client.DispatchEx("Excel.Application")
            wb = xl.Workbooks.Open(os.path.abspath(file_name))
            
            try:
                for ws in wb.Worksheets:
                    ws.Columns("A:A").EntireColumn.AutoFit()
                wb.Save()
            finally:
                wb.Close(SaveChanges=True)
                xl.Quit()
            
            settings = QSettings('MyApp', 'ExcelExport')
            settings.setValue('last_export_path', file_name)
            
            QMessageBox.information(self, "Sucesso", "Arquivo Excel exportado com sucesso!")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar planilha:\n{str(e)}")

    def clean_sheet_name(self, name):
        """Limpa e valida o nome da planilha"""
        # Caracteres inválidos em nomes de planilhas do Excel
        invalid_chars = ['/', '\\', '?', '*', ':', '[', ']']
        
        # Substituir caracteres inválidos por underscore
        clean_name = name
        for char in invalid_chars:
            clean_name = clean_name.replace(char, '_')
        
        # Limitar o tamanho do nome (Excel permite até 31 caracteres)
        clean_name = clean_name[:31]
        
        # Se o nome começar com número, adicionar prefixo
        if clean_name[0].isdigit():
            clean_name = 'S_' + clean_name
        
        # Garantir que o nome não está vazio
        if not clean_name:
            clean_name = 'Sheet'
        
        return clean_name

    def add_frase(self):
        text = self.filter_input.text()
        if text and self.active_frases is not None:
            self.active_frases.append(text)
            self.active_frases = sorted(self.active_frases)
            self.frases[self.active_button.text()] = self.active_frases
            self.mostrar_frases()
            self.filter_input.clear()
            self.save_phrases_to_file()

    def remove_frase(self):
        selected_items = self.messages_list.selectedItems()
        if selected_items:
            for item in selected_items:
                self.active_frases.remove(item.text())
            self.active_frases = sorted(self.active_frases)
            self.frases[self.active_button.text()] = self.active_frases
            self.mostrar_frases()
            self.save_phrases_to_file()

    def mostrar_frases(self):
        self.messages_list.clear()
        for frase in sorted(self.active_frases):
            item = QListWidgetItem(frase)
            self.messages_list.addItem(item)

    def show_context_menu(self, position):
        menu = QMenu(self)
        remove_action = menu.addAction("Remover Frase")
        action = menu.exec_(self.messages_list.mapToGlobal(position))
        if action == remove_action:
            self.remove_frase()

    def load_txt_file(self, auto_load=False):
        """Carrega um arquivo TXT para visualização"""
        try:
            # Carregar último diretório e arquivo usado
            settings = QSettings('MyApp', 'TextViewer')
            last_directory = settings.value('last_txt_directory', '')
            last_file = settings.value('last_txt_file', '')
            
            if auto_load and last_file and os.path.exists(last_file):
                file_name = last_file
            else:
                file_name, _ = QFileDialog.getOpenFileName(
                    self,
                    "Abrir Arquivo TXT",
                    last_directory,
                    "Arquivos de Texto (*.txt);;Todos os Arquivos (*)"
                )
            
            if file_name:
                settings.setValue('last_txt_directory', os.path.dirname(file_name))
                settings.setValue('last_txt_file', file_name)
                
                with open(file_name, 'r', encoding='utf-8') as file:
                    content = file.read()
                    
                    main_widget = QWidget()
                    main_layout = QVBoxLayout(main_widget)
                    main_layout.setSpacing(2)
                    main_layout.setContentsMargins(5, 5, 5, 5)
                    
                    scroll = QScrollArea()
                    scroll.setWidgetResizable(True)
                    
                    content_widget = QWidget()
                    content_layout = QVBoxLayout(content_widget)
                    content_layout.setSpacing(1)
                    content_layout.setContentsMargins(2, 2, 2, 2)
                    
                    self.line_widgets = []
                    
                    lines = content.split('\n')
                    for line in lines:
                        if line.strip():
                            line_widget = QWidget()
                            line_layout = QHBoxLayout(line_widget)
                            line_layout.setContentsMargins(0, 0, 0, 0)
                            line_layout.setSpacing(5)
                            
                            checkbox = QCheckBox()
                            checkbox.setFixedSize(15, 15)
                            line_layout.addWidget(checkbox)
                            
                            text_label = QLabel(line)
                            text_label.setWordWrap(True)
                            text_label.setFont(QFont("Arial", 9))
                            text_label.setStyleSheet("""
                                QLabel {
                                    padding: 1px;
                                    background-color: transparent;
                                }
                            """)
                            text_label.setTextInteractionFlags(
                                Qt.TextInteractionFlag.TextSelectableByMouse | 
                                Qt.TextInteractionFlag.TextSelectableByKeyboard
                            )
                            line_layout.addWidget(text_label, 1)
                            
                            checkbox.stateChanged.connect(
                                lambda state, label=text_label: self.handle_checkbox(state, label))
                            
                            content_layout.addWidget(line_widget)
                            self.line_widgets.append((checkbox, text_label))
                    
                    content_layout.addStretch()
                    scroll.setWidget(content_widget)
                    main_layout.addWidget(scroll)
                    
                    # Botões de controle
                    control_bar = QWidget()
                    control_layout = QHBoxLayout(control_bar)
                    control_layout.setContentsMargins(0, 0, 0, 0)
                    
                    select_all_btn = QPushButton("Selecionar Todos")
                    deselect_all_btn = QPushButton("Desselecionar Todos")
                    
                    select_all_btn.setFixedHeight(25)
                    deselect_all_btn.setFixedHeight(25)
                    
                    control_layout.addWidget(select_all_btn)
                    control_layout.addWidget(deselect_all_btn)
                    control_layout.addStretch()
                    
                    select_all_btn.clicked.connect(self.select_all_lines)
                    deselect_all_btn.clicked.connect(self.deselect_all_lines)
                    
                    main_layout.addWidget(control_bar)
                    
                    if self.text_viewer.layout():
                        while self.text_viewer.layout().count():
                            item = self.text_viewer.layout().takeAt(0)
                            if item.widget():
                                item.widget().deleteLater()
                    
                    if not self.text_viewer.layout():
                        self.text_viewer.setLayout(QVBoxLayout())
                    
                    self.text_viewer.layout().addWidget(main_widget)
                    
                    file_size = os.path.getsize(file_name) / 1024
                    file_name_short = os.path.basename(file_name)
                    self.status_bar.showMessage(
                        f"Arquivo carregado: {file_name_short} ({file_size:.1f} KB)",
                        3000
                    )
                    
        except Exception as e:
            if not auto_load:  # Só mostra erro se não for carregamento automático
                QMessageBox.warning(self, "Erro", f"Erro ao carregar arquivo: {str(e)}")

    def handle_checkbox(self, state, label):
        """Gerencia a seleção via checkbox silenciosamente"""
        if state == Qt.CheckState.Checked:
            label.setStyleSheet("""
                QLabel {
                    background-color: #e6f3ff;
                    border-radius: 2px;
                    padding: 1px;
                }
            """)
            if label.text() not in self.selected_texts:
                self.selected_texts.append(label.text())
        else:
            label.setStyleSheet("""
                QLabel {
                    background-color: transparent;
                    padding: 1px;
                }
            """)
            if label.text() in self.selected_texts:
                self.selected_texts.remove(label.text())

    def select_all_lines(self):
        """Seleciona todas as linhas"""
        try:
            for checkbox, _ in self.line_widgets:
                checkbox.setChecked(True)
        except Exception as e:
            print(f"Erro ao selecionar todas as linhas: {str(e)}")

    def deselect_all_lines(self):
        """Desmarca todas as linhas"""
        try:
            for checkbox, _ in self.line_widgets:
                checkbox.setChecked(False)
        except Exception as e:
            print(f"Erro ao desmarcar todas as linhas: {str(e)}")

    def show_copy_dialog(self):
        """Mostra diálogo para selecionar destino da cópia"""
        # Coletar textos dos checkboxes marcados
        selected_texts = []
        for checkbox, label in self.line_widgets:
            if checkbox.isChecked():
                selected_texts.append(label.text())
        
        if not selected_texts:  # Verifica se há textos selecionados via checkbox
            QMessageBox.warning(self, "Aviso", "Selecione pelo menos um texto usando os checkboxes!")
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Selecionar Destino")
        dialog.setMinimumWidth(400)
        layout = QVBoxLayout()
        
        # Scroll area para os segmentos e botões
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        checkboxes = {}
        
        # Criar estrutura de segmentos e botões
        for segmento, botoes in self.segmentos.items():
            segment_group = QGroupBox(segmento)
            segment_layout = QVBoxLayout()
            
            for botao in botoes:
                button_layout = QHBoxLayout()
                checkbox = QCheckBox(botao)
                checkbox.setFont(QFont("Arial", 9))
                button_layout.addWidget(checkbox)
                button_layout.addStretch()
                
                segment_layout.addLayout(button_layout)
                checkboxes[botao] = checkbox
            
            segment_group.setLayout(segment_layout)
            scroll_layout.addWidget(segment_group)
        
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)
        
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | 
            QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        dialog.setLayout(layout)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            copied = False
            for botao, checkbox in checkboxes.items():
                if checkbox.isChecked():
                    # Garantir que o botão existe no dicionário de frases
                    if botao not in self.frases:
                        self.frases[botao] = []
                    
                    # Adicionar novas frases evitando duplicatas
                    for texto in selected_texts:  # Usa os textos coletados dos checkboxes
                        if texto not in self.frases[botao]:
                            self.frases[botao].append(texto)
                    
                    # Ordenar as frases
                    self.frases[botao].sort()
                    copied = True
                    
                    # Se este for o botão ativo, atualizar a visualização
                    if self.active_button and self.active_button.text() == botao:
                        self.active_frases = self.frases[botao]
                        self.mostrar_frases()
            
            if copied:
                # Salvar as alterações no arquivo
                if self.save_phrases_to_file():
                    QMessageBox.information(
                        self,
                        "Sucesso",
                        "Frases copiadas com sucesso!"
                    )
                else:
                    QMessageBox.warning(
                        self,
                        "Aviso",
                        "As frases foram copiadas, mas houve um erro ao salvar o arquivo!"
                    )

    def eventFilter(self, obj, event):
        if obj is self.text_viewer and event.type() == QEvent.KeyPress:
            if event.key() == Qt.Key_Control:
                # Habilita seleção múltipla quando CTRL é pressionado
                self.text_viewer.setTextInteractionFlags(
                    Qt.TextSelectableByMouse | Qt.TextSelectableByKeyboard
                )
                return True
        elif obj is self.text_viewer and event.type() == QEvent.KeyRelease:
            if event.key() == Qt.Key_Control:
                # Desabilita seleção múltipla quando CTRL é liberado
                self.text_viewer.setTextInteractionFlags(Qt.TextSelectableByMouse)
                return True
        return super().eventFilter(obj, event)

    def setup_text_viewer(self):
        """Configura o visualizador de texto com suporte a múltiplas seleções"""
        self.text_viewer = QTextEdit()
        self.text_viewer.setReadOnly(True)
        self.text_viewer.installEventFilter(self)
        self.text_viewer.setTextInteractionFlags(Qt.TextSelectableByMouse)

    def custom_mouse_press_event(self, event):
        """Manipula o evento de pressionar o mouse no visualizador de texto"""
        try:
            if event.button() == Qt.LeftButton:
                cursor = self.text_viewer.cursorForPosition(event.pos())
                if event.modifiers() & Qt.ControlModifier:
                    # Seleção com CTRL pressionado
                    cursor.movePosition(QTextCursor.StartOfLine)
                    cursor.movePosition(QTextCursor.EndOfLine, QTextCursor.KeepAnchor)
                    selected_text = cursor.selectedText()
                    
                    if selected_text.strip():
                        if selected_text not in self.selected_texts:
                            self.selected_texts.append(selected_text)
                            
                            # Destaca a seleção
                            extra_selection = QTextEdit.ExtraSelection()
                            extra_selection.format.setBackground(QColor(173, 216, 230))
                            extra_selection.cursor = cursor
                            
                            current_selections = self.text_viewer.extraSelections()
                            current_selections.append(extra_selection)
                            self.text_viewer.setExtraSelections(current_selections)
                else:
                    # Seleção normal
                    QTextEdit.mousePressEvent(self.text_viewer, event)
                    self.selected_texts.clear()
                    self.text_viewer.setExtraSelections([])
        except Exception as e:
            print(f"Erro no evento de mouse: {str(e)}")

    def custom_mouse_release_event(self, event):
        """Manipula o evento de soltar o botão do mouse"""
        try:
            if event.button() == Qt.LeftButton and not (event.modifiers() & Qt.ControlModifier):
                cursor = self.text_viewer.textCursor()
                if cursor.hasSelection():
                    selected_text = cursor.selectedText()
                    if selected_text.strip():
                        self.selected_texts = [selected_text]
            QTextEdit.mouseReleaseEvent(self.text_viewer, event)
        except Exception as e:
            print(f"Erro no evento de soltar mouse: {str(e)}")

    def open_export_folder(self):
        """Abre o diretório de exportação dos arquivos Excel"""
        try:
            settings = QSettings('MyApp', 'ExcelExport')
            last_export_path = settings.value('last_export_path')
            
            # Define o diretório padrão como 'Downloads' se não houver último caminho
            if last_export_path and os.path.exists(os.path.dirname(last_export_path)):
                target_dir = os.path.dirname(last_export_path)
            else:
                # Tenta usar o diretório Downloads
                if os.name == 'nt':  # Windows
                    target_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
                else:  # Linux/Mac
                    target_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
                
                # Se Downloads no existir, usa o diretório do usuário
                if not os.path.exists(target_dir):
                    target_dir = os.path.expanduser('~')
            
            # Abre o explorador de arquivos no diretório
            if os.name == 'nt':  # Windows
                os.startfile(target_dir)
            else:  # Linux/Mac
                import subprocess
                subprocess.Popen(['xdg-open', target_dir])
            
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao abrir o diretório: {str(e)}")

    def get_greeting(self):
        """Retorna a saudação apropriada baseada no horário atual"""
        from datetime import datetime
        hora_atual = datetime.now().hour
        minuto_atual = datetime.now().minute
        
        # Converte hora e minuto para minutos totais para facilitar a comparação
        tempo_atual = hora_atual * 60 + minuto_atual
        
        # Define os limites dos turnos em minutos
        manha_inicio = 8 * 60 + 30  # 08:30
        manha_fim = 11 * 60 + 59    # 11:59
        tarde_fim = 17 * 60 + 59    # 17:59
        noite_fim = 23 * 60 + 59    # 23:59
        
        if manha_inicio <= tempo_atual <= manha_fim:
            return "Bom dia!"
        elif tempo_atual <= tarde_fim:
            return "Boa tarde!"
        elif tempo_atual <= noite_fim:
            return "Boa noite!"
        else:
            return "Olá!"  # Fora dos horários especificados

    def open_outlook_email(self, search_term):
        try:
            import win32com.client
            outlook = win32com.client.Dispatch("Outlook.Application")
            namespace = outlook.GetNamespace("MAPI")
            
            # Acessar a caixa de saída
            sent_folder = namespace.GetDefaultFolder(5)  # 5 = olFolderSentMail
            
            # Procurar pelo e-mail mais recente com o termo especificado
            items = sent_folder.Items
            items.Sort("[ReceivedTime]", True)  # Ordenar por data, mais recente primeiro
            
            found_email = None
            for item in items:
                if search_term in item.Subject:
                    found_email = item
                    break
            
            if found_email:
                # Criar uma nova mensagem baseada no e-mail encontrado
                new_mail = outlook.CreateItem(0)  # 0 = olMailItem
                new_mail.Subject = found_email.Subject
                
                # Processar o corpo do e-mail baseado no checkbox
                if self.remove_images_cb.isChecked():
                    # Remove tags de imagem do HTML
                    html_body = found_email.HTMLBody
                    import re
                    html_body = re.sub(r'<img[^>]*>', '', html_body)
                    html_body = re.sub(r'<v:imagedata[^>]*>', '', html_body)
                    html_body = re.sub(r'<v:image[^>]*>.*?</v:image>', '', html_body)
                    new_mail.HTMLBody = html_body
                else:
                    new_mail.HTMLBody = found_email.HTMLBody
                
                # Substituir destinatário baseado nos checkboxes de ausência
                original_to = found_email.To
                if search_term == "Alerta 8" and self.evandro_absent_cb.isChecked():
                    original_to = original_to.replace(
                        "evandro.britto@bradescoseguros.com.br",
                        "aline.campos@bradescoseguros.com.br"
                    )
                elif search_term == "Alerta 9" and self.aline_absent_cb.isChecked():
                    original_to = original_to.replace(
                        "aline.campos@bradescoseguros.com.br",
                        "evandro.britto@bradescoseguros.com.br"
                    )
                
                new_mail.To = original_to
                
                # Adicionar saudação apropriada no início do corpo do e-mail
                greeting = self.get_greeting()
                new_mail.HTMLBody = f"{greeting}<br><br>" + new_mail.HTMLBody
                
                # Copiar anexos se o checkbox não estiver marcado
                if not self.remove_attachments_cb.isChecked():
                    for attachment in found_email.Attachments:
                        new_mail.Attachments.Add(attachment.PropertyAccessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F"))
                
                new_mail.Display()  # Mostrar o novo e-mail
            else:
                QMessageBox.warning(self, "Aviso", f"Nenhum e-mail encontrado com o termo '{search_term}'")
                
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao abrir o Outlook: {str(e)}")

    def add_site_thumbnail(self, site_data):
        """Adiciona thumbnail do site com preview estático"""
        try:
            # Frame principal do thumbnail
            site_frame = QFrame()
            # Corrigir as flags do QFrame para PyQt6
            site_frame.setFrameStyle(QFrame.Shape.Box | QFrame.Shadow.Plain)
            site_frame.setLineWidth(1)
            frame_layout = QVBoxLayout()
            
            # Label com nome do site
            site_name_label = QLabel(site_data['name'])
            site_name_label.setFixedWidth(250)
            site_name_label.setMinimumHeight(25)
            site_name_label.setMaximumHeight(25)
            site_name_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            site_name_label.setStyleSheet("""
                QLabel {
                    color: #333333;
                    font-size: 11px;
                    font-weight: bold;
                    padding: 2px;
                    margin: 0px;
                    background-color: #f5f5f5;
                    border: 1px solid #ddd;
                    border-radius: 3px;
                }
            """)
            
            # Frame de login
            login_frame = QFrame()
            # Corrigir as flags do QFrame para PyQt6
            login_frame.setFrameStyle(QFrame.Shape.Box | QFrame.Shadow.Plain)
            login_frame.setLineWidth(1)
            login_layout = QGridLayout()
            
            # Campos de login
            login_label = QLabel("Login:")
            username_label = QLabel("Usuário:")
            password_label = QLabel("Senha:")
            username_field = QLineEdit(site_data['username'])
            password_field = QLineEdit(site_data['password'])
            # Corrigir o modo de senha para PyQt6
            password_field.setEchoMode(QLineEdit.EchoMode.Password)
            
            login_btn = QPushButton("Logar")
            keep_active_cb = QCheckBox("Manter conexão ativa")
            
            # Adicionar widgets ao layout de login
            login_layout.addWidget(login_label, 0, 0)
            login_layout.addWidget(username_label, 1, 0)
            login_layout.addWidget(username_field, 1, 1)
            login_layout.addWidget(password_label, 2, 0)
            login_layout.addWidget(password_field, 2, 1)
            login_layout.addWidget(login_btn, 3, 0)
            login_layout.addWidget(keep_active_cb, 3, 1)
            
            login_frame.setLayout(login_layout)
            
            # Montar layout do thumbnail
            frame_layout.addWidget(site_name_label)
            frame_layout.addWidget(login_frame)
            frame_layout.setSpacing(5)
            frame_layout.setContentsMargins(10, 5, 10, 10)
            site_frame.setLayout(frame_layout)
            
            # Adicionar à grid
            col = self.sites_grid.count() % 3
            row = self.sites_grid.count() // 3
            self.sites_grid.addWidget(site_frame, row, col)
            
            # Capturar preview do site
            self.capture_site_preview(site_data['url'], site_name_label)
            
            return True
            
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao criar thumbnail: {str(e)}")
            return False

    def capture_site_preview(self, url, preview_label):
        """Captura o preview do site usando QWebEngineView"""
        try:
            view = QWebEngineView()
            view.setFixedSize(1024, 768)  # Tamanho maior para melhor qualidade
            
            # Configurações adicionais para melhor renderização
            page = view.page()
            settings = page.settings()
            settings.setAttribute(QWebEngineSettings.WebAttribute.ShowScrollBars, False)
            settings.setAttribute(QWebEngineSettings.WebAttribute.JavascriptEnabled, True)
            
            def handle_load_finished(success):
                if success:
                    # Aguardar um pouco para carregar conteúdo dinâmico
                    QTimer.singleShot(2000, take_snapshot)
                else:
                    preview_label.setMovie(None)  # Remove loading
                    preview_label.setText("Erro ao carregar preview")
                    view.deleteLater()
            
            def take_snapshot():
                try:
                    # Capturar página com melhor qualidade
                    page.runJavaScript("""
                        document.documentElement.style.overflow = 'hidden';
                        document.body.style.overflow = 'hidden';
                    """)
                    
                    view.grab().then(lambda image: process_image(image))
                except Exception as e:
                    preview_label.setMovie(None)
                    preview_label.setText("Erro na captura")
                    print(f"Erro: {str(e)}")
                finally:
                    view.deleteLater()
            
            def process_image(pixmap):
                try:
                    # Aplicar efeitos para melhor aparência
                    scaled = pixmap.scaled(
                        250, 180,
                        Qt.AspectRatioMode.KeepAspectRatioByExpanding,
                        Qt.TransformationMode.SmoothTransformation
                    )
                    
                    # Adicionar sombra e borda
                    final_pixmap = QPixmap(scaled.size())
                    final_pixmap.fill(Qt.GlobalColor.transparent)
                    
                    painter = QPainter(final_pixmap)
                    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                    painter.drawPixmap(0, 0, scaled)
                    painter.end()
                    
                    preview_label.setMovie(None)  # Remove loading
                    preview_label.setPixmap(final_pixmap)
                except Exception as e:
                    preview_label.setText("Erro no processamento")
                    print(f"Erro: {str(e)}")

            # Iniciar carregamento
            view.loadFinished.connect(handle_load_finished)
            view.load(QUrl(url))
            
        except Exception as e:
            preview_label.setText("Erro na inicialização")
            print(f"Erro: {str(e)}")

    def open_site_in_tab(self, site_data):
        """Abre o site em uma nova aba"""
        try:
            web_view = QWebEngineView()
            web_view.setUrl(QUrl(site_data['url']))
            
            # Adicionar nova aba
            index = self.sites_tabs.addTab(web_view, site_data['name'])
            self.sites_tabs.setCurrentIndex(index)
            
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao abrir site: {str(e)}")

    def remove_site_dialog(self):
        """Diálogo para remover sites"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Remover Sites")
        dialog.setMinimumWidth(400)
        layout = QVBoxLayout()

        # Lista de sites com checkboxes
        scroll = QScrollArea()
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        
        # Lista para armazenar todos os checkboxes e seus dados associados
        all_checkboxes = []
        sites_data = []
        
        # Coletar todos os sites da grid
        for i in range(self.sites_grid.count()):
            item = self.sites_grid.itemAt(i)
            if item and item.widget():
                site_frame = item.widget()
                site_btn = site_frame.findChild(QPushButton)
                if site_btn:
                    site_name = site_btn.text()
                    cb = QCheckBox(site_name)
                    all_checkboxes.append(cb)
                    
                    # Armazenar dados do site
                    sites_data.append({
                        'checkbox': cb,
                        'frame': site_frame,
                        'name': site_name,
                        'index': i
                    })
                    
                    scroll_layout.addWidget(cb)
        
        scroll_content.setLayout(scroll_layout)
        scroll.setWidget(scroll_content)
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        # Botões de seleção
        select_buttons = QHBoxLayout()
        select_all = QPushButton("Selecionar Todos")
        deselect_all = QPushButton("Desselecionar Todos")
        
        def select_all_boxes():
            for checkbox in all_checkboxes:
                checkbox.setChecked(True)
        
        def deselect_all_boxes():
            for checkbox in all_checkboxes:
                checkbox.setChecked(False)
        
        select_all.clicked.connect(select_all_boxes)
        deselect_all.clicked.connect(deselect_all_boxes)
        
        select_buttons.addWidget(select_all)
        select_buttons.addWidget(deselect_all)
        
        # Botões OK/Cancelar
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        
        layout.addWidget(scroll)
        layout.addLayout(select_buttons)
        layout.addWidget(buttons)
        dialog.setLayout(layout)
        
        if dialog.exec_() == QDialog.Accepted:
            # Identificar sites selecionados para remoção
            sites_to_remove = [
                site_data for site_data in sites_data
                if site_data['checkbox'].isChecked()
            ]
            
            if sites_to_remove:
                # Confirmar remoção
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Question)
                msg.setText(f"Deseja remover {len(sites_to_remove)} site(s)?")
                msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                msg.setDefaultButton(QMessageBox.No)
                
                if msg.exec_() == QMessageBox.Yes:
                    # Remover abas primeiro
                    for site_data in sites_to_remove:
                        site_name = site_data['name']
                        # Remover abas associadas (de trás para frente)
                        for i in range(self.sites_tabs.count() - 1, 0, -1):
                            if self.sites_tabs.tabText(i) == site_name:
                                self.sites_tabs.removeTab(i)
                
                    # Remover widgets da grid (de trás para frente)
                    for site_data in reversed(sites_to_remove):
                        frame = site_data['frame']
                        self.sites_grid.removeWidget(frame)
                        frame.deleteLater()
                    
                    # Reorganizar grid
                    self.reorganize_grid()
                    
                    # Salvar alterações
                    self.save_sites()
                    
                    QMessageBox.information(
                        dialog,
                        "Sucesso",
                        f"{len(sites_to_remove)} site(s) removido(s) com sucesso!"
                    )

    def close_site_tab(self, index):
        """Fecha uma aba de site, exceto a aba de relação"""
        if index > 0:  # Não fecha a aba de relaço
            self.sites_tabs.removeTab(index)

    def save_sites(self):
        """Salva os dados dos sites em um arquivo JSON"""
        try:
            sites_data = []
            for i in range(self.sites_grid.count()):
                item = self.sites_grid.itemAt(i)
                if item and item.widget():
                    frame = item.widget()
                    # Extrair dados do frame
                    # Implementar lógica para extrair dados
                    sites_data.append({
                        'name': 'Site Name',
                        'url': 'URL',
                        'username': 'Username',
                        'password': 'Password'  # Considerar criptografia
                    })
            
            with open('sites.json', 'w') as f:
                json.dump(sites_data, f)
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao salvar sites: {str(e)}")

    def load_saved_sites(self):
        """Carrega os sites salvos do arquivo JSON"""
        try:
            if os.path.exists('sites.json'):
                with open('sites.json', 'r') as f:
                    sites_data = json.load(f)
                    for site_data in sites_data:
                        self.add_site_thumbnail(site_data)
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao carregar sites: {str(e)}")

    def reorganize_grid(self):
        """Reorganiza os thumbnails na grid após remoções"""
        # Coletar todos os widgets restantes
        widgets = []
        while self.sites_grid.count() > 0:
            item = self.sites_grid.itemAt(0)
            if item and item.widget():
                widget = item.widget()
                self.sites_grid.removeWidget(widget)
                if not widget.isHidden() and widget.isVisible():
                    widgets.append(widget)
        
        # Readicionar os widgets à grid
        for i, widget in enumerate(widgets):
            row = i // 3
            col = i % 3
            self.sites_grid.addWidget(widget, row, col)
            widget.show()  # Garantir que o widget está visível
        
        # Forçar atualização do layout
        self.sites_grid.update()

    def closeEvent(self, event):
        # Limpar recursos do WebEngine antes de fechar
        for i in range(self.tab_widget.count()):
            widget = self.tab_widget.widget(i)
            if isinstance(widget, QWebEngineView):
                widget.setParent(None)
        event.accept()

    def update_excel(self):
        """Atualiza o último arquivo Excel salvo com as frases atuais"""
        try:
            settings = QSettings('MyApp', 'ExcelExport')
            last_export_path = settings.value('last_export_path')
            
            if not last_export_path or not os.path.exists(last_export_path):
                QMessageBox.warning(self, "Aviso", "Nenhum arquivo Excel encontrado. Por favor, exporte primeiro.")
                return

            workbook = Workbook()
            
            for segmento, botoes in self.segmentos.items():
                for botao in botoes:
                    if botao in self.frases and self.frases[botao]:
                        sheet_name = botao[:31].upper()  # Nome em maiúsculas
                        if sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                        else:
                            sheet = workbook.create_sheet(title=sheet_name)
                        
                        # Título em maiúsculas
                        header_cell = sheet.cell(row=1, column=1, value=botao.upper())
                        header_cell.font = Font(name='Calibri', size=8, bold=True)
                        header_cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Calcular largura inicial
                        max_length = len(botao.upper())
                        for frase in self.frases[botao]:
                            max_length = max(max_length, len(str(frase)))
                        sheet.column_dimensions['A'].width = (max_length + 2) * 1.2
                        
                        for idx, frase in enumerate(sorted(self.frases[botao]), start=2):
                            cell = sheet.cell(row=idx, column=1, value=str(frase))
                            cell.font = Font(name='Calibri', size=8)
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        for row in sheet.rows:
                            sheet.row_dimensions[row[0].row].height = 15

            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
                
            # Salvar arquivo
            workbook.save(last_export_path)
            
            # Aplicar AutoFit usando win32com
            import win32com.client
            xl = win32com.client.DispatchEx("Excel.Application")
            wb = xl.Workbooks.Open(os.path.abspath(last_export_path))
            
            try:
                for ws in wb.Worksheets:
                    ws.Columns("A:A").EntireColumn.AutoFit()
                wb.Save()
            finally:
                wb.Close(SaveChanges=True)
                xl.Quit()
            
            QMessageBox.information(self, "Sucesso", "Planilha atualizada com sucesso!")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao atualizar planilha:\n{str(e)}")

    def open_new_email(self):
        """Abre uma nova janela de e-mail no Outlook"""
        try:
            import win32com.client
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)  # 0 = olMailItem
            mail.Display()
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao abrir o Outlook: {str(e)}")

    def add_site_to_list(self):
        """Mostra diálogo para adicionar site"""
        dialogo = QDialog(self)
        dialogo.setWindowTitle("Adicionar Site")
        layout = QVBoxLayout()

        # Campos básicos
        nome_label = QLabel("Nome do Site:")
        nome_input = QLineEdit()
        url_label = QLabel("URL:")
        url_input = QLineEdit()

        # Checkbox para credenciais
        tem_login = QCheckBox("Site requer login")
        
        # Campos de login (inicialmente ocultos)
        login_widget = QWidget()
        login_layout = QVBoxLayout(login_widget)
        
        login_label = QLabel("Login:")
        login_input = QLineEdit()
        senha_label = QLabel("Senha:")
        senha_input = QLineEdit()
        senha_input.setEchoMode(QLineEdit.EchoMode.Password)
        
        login_layout.addWidget(login_label)
        login_layout.addWidget(login_input)
        login_layout.addWidget(senha_label)
        login_layout.addWidget(senha_input)
        login_widget.hide()

        # Conectar checkbox aos campos de login
        def toggle_login_fields(state):
            login_widget.setVisible(state == Qt.CheckState.Checked.value)
            dialogo.adjustSize()

        tem_login.stateChanged.connect(toggle_login_fields)

        # Botões
        botoes = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | 
            QDialogButtonBox.StandardButton.Cancel
        )

        # Adicionar widgets ao layout
        layout.addWidget(nome_label)
        layout.addWidget(nome_input)
        layout.addWidget(url_label)
        layout.addWidget(url_input)
        layout.addWidget(tem_login)
        layout.addWidget(login_widget)
        layout.addWidget(botoes)

        dialogo.setLayout(layout)

        def salvar_site():
            nome = nome_input.text().strip()
            url = url_input.text().strip()
            
            if not url.startswith(('http://', 'https://')):
                url = 'https://' + url
                
            if nome and url:
                try:
                    # Carregar sites existentes
                    try:
                        with open('sites.json', 'r', encoding='utf-8') as f:
                            sites = json.load(f)
                            # Verificar se a URL já existe
                            for site in sites:
                                if site['url'].lower().strip('/') == url.lower().strip('/'):
                                    QMessageBox.warning(dialogo, "Erro", "Este endereço web já está cadastrado!")
                                    return
                    except FileNotFoundError:
                        sites = []

                    # Criar novo site
                    novo_site = {
                        'nome': nome,
                        'url': url,
                        'login': login_input.text() if tem_login.isChecked() else None,
                        'senha': senha_input.text() if tem_login.isChecked() else None
                    }
                    
                    # Adicionar à lista de sites
                    sites.append(novo_site)

                    # Salvar no arquivo
                    with open('sites.json', 'w', encoding='utf-8') as f:
                        json.dump(sites, f, ensure_ascii=False, indent=4)

                    # Adicionar à lista visual
                    item = QListWidgetItem(nome)
                    self.sites_list.addItem(item)
                    self.sites_list.setCurrentItem(item)
                    self.sites_list.setFocus()

                    QMessageBox.information(dialogo, "Sucesso", "Site adicionado com sucesso!")
                    dialogo.accept()
                    
                except Exception as e:
                    QMessageBox.warning(dialogo, "Erro", f"Erro ao adicionar site: {str(e)}")
            else:
                QMessageBox.warning(dialogo, "Erro", "Nome e URL são obrigatórios!")

        botoes.accepted.connect(salvar_site)
        botoes.rejected.connect(dialogo.reject)

        dialogo.exec()

    def remove_site_from_list(self):
        """Remove sites selecionados da lista"""
        selected_rows = set(item.row() for item in self.sites_table.selectedItems())
        if not selected_rows:
            QMessageBox.warning(self, "Aviso", "Selecione pelo menos um site para remover.")
            return
        
        reply = QMessageBox.question(
            self, 
            'Confirmar Remoção',
            f'Deseja remover {len(selected_rows)} site(s)?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            for row in sorted(selected_rows, reverse=True):
                self.sites_table.removeRow(row)
            self.save_sites_list()

    def export_sites_list(self):
        """Exporta a lista de sites para Excel"""
        try:
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Exportar Lista de Sites",
                "",
                "Excel Files (*.xlsx)"
            )
            
            if file_name:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Sites"
                
                # Cabeçalhos
                headers = ["Nome", "URL", "Usuário", "Senha"]
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col, value=header)
                
                # Dados
                for row in range(self.sites_table.rowCount()):
                    for col in range(self.sites_table.columnCount()):
                        item = self.sites_table.item(row, col)
                        value = item.text() if item else ""
                        sheet.cell(row=row+2, column=col+1, value=value)
                
                workbook.save(file_name)
                QMessageBox.information(self, "Sucesso", "Lista exportada com sucesso!")
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar: {str(e)}")

    def save_sites_list(self):
        """Salva a lista de sites em um arquivo JSON"""
        try:
            sites_data = []
            for row in range(self.sites_table.rowCount()):
                site = {
                    'name': self.sites_table.item(row, 0).text(),
                    'url': self.sites_table.item(row, 1).text(),
                    'username': self.sites_table.item(row, 2).text(),
                    'password': self.sites_table.item(row, 3).text()
                }
                sites_data.append(site)
                
            with open('sites_list.json', 'w', encoding='utf-8') as f:
                json.dump(sites_data, f, ensure_ascii=False, indent=4)
                
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao salvar lista: {str(e)}")

    def load_sites_list(self):
        """Carrega a lista de sites do arquivo JSON"""
        try:
            if os.path.exists('sites_list.json'):
                with open('sites_list.json', 'r', encoding='utf-8') as f:
                    sites_data = json.load(f)
                    
                    self.sites_table.setRowCount(0)
                    for site in sites_data:
                        row = self.sites_table.rowCount()
                        self.sites_table.insertRow(row)
                        self.sites_table.setItem(row, 0, QTableWidgetItem(site['name']))
                        self.sites_table.setItem(row, 1, QTableWidgetItem(site['url']))
                        self.sites_table.setItem(row, 2, QTableWidgetItem(site['username']))
                        self.sites_table.setItem(row, 3, QTableWidgetItem(site['password']))
                        
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao carregar lista: {str(e)}")

    def on_site_selected(self, item):
        """Callback quando um site é selecionado na lista"""
        try:
            if item:
                site_info = item.text()
                
                # Criar container para a aba
                container = QWidget()
                layout = QVBoxLayout(container)
                layout.setContentsMargins(0, 0, 0, 0)  # Remove todas as margens
                layout.setSpacing(0)  # Remove espaçamento entre elementos
                
                # Criar barra de navegação
                nav_bar = QHBoxLayout()
                nav_bar.setContentsMargins(1, 0, 1, 1)  # Reduz ainda mais as margens, especialmente no topo
                nav_bar.setSpacing(1)  # Reduz espaçamento entre botões
                
                # Botões de navegação
                back_button = QPushButton("Voltar")
                back_button.setMaximumWidth(60)
                forward_button = QPushButton("Avançar")
                forward_button.setMaximumWidth(60)
                refresh_button = QPushButton("Atualizar")
                refresh_button.setMaximumWidth(60)
                
                # Campo de endereço
                address_bar = QLineEdit()
                address_bar.setPlaceholderText("Digite o endereço da web")
                
                # Adicionar widgets à barra de navegação
                nav_bar.addWidget(back_button)
                nav_bar.addWidget(forward_button)
                nav_bar.addWidget(refresh_button)
                nav_bar.addWidget(address_bar)
                
                # Criar widget para conter a barra de navegação
                nav_widget = QWidget()
                nav_widget.setLayout(nav_bar)
                nav_widget.setContentsMargins(0, 0, 0, 0)  # Remove todas as margens do widget de navegação
                nav_widget.setFixedHeight(30)  # Define altura fixa para a barra de navegação
                
                # Carregar sites do arquivo JSON
                with open('sites.json', 'r', encoding='utf-8') as f:
                    sites = json.load(f)
                
                # Encontrar o site selecionado
                site = next((s for s in sites if s['nome'] == site_info), None)
                
                if site:
                    url = site['url']
                    if not url.startswith(('http://', 'https://')):
                        url = 'https://' + url
                    
                    # Criar visualizador web
                    web_view = QWebEngineView()
                    web_view.setUrl(QUrl(url))
                    web_view.setContentsMargins(0, 0, 0, 0)
                    
                    # Conectar botões e campo de endereço
                    back_button.clicked.connect(web_view.back)
                    forward_button.clicked.connect(web_view.forward)
                    refresh_button.clicked.connect(web_view.reload)
                    address_bar.returnPressed.connect(
                        lambda: self.load_url(web_view, address_bar.text())
                    )
                    
                    # Atualizar o campo de endereço quando a URL mudar
                    web_view.urlChanged.connect(lambda qurl: address_bar.setText(qurl.toString()))
                    
                    # Adicionar widgets ao layout
                    layout.addWidget(nav_widget)
                    layout.addWidget(web_view)
                    
                    # Verificar se já existe uma aba com este site
                    for i in range(self.sites_tabs.count()):
                        if self.sites_tabs.tabText(i) == site_info:
                            self.sites_tabs.setCurrentIndex(i)
                            return
                    
                    # Adicionar nova aba
                    index = self.sites_tabs.addTab(container, site_info)
                    self.sites_tabs.setCurrentIndex(index)
                    
        except Exception as e:
            print(f"Erro ao carregar site: {str(e)}")
            QMessageBox.warning(
                self,
                "Erro",
                f"Erro ao carregar site: {str(e)}"
            )

    def update_tab_title(self, browser, title):
        """Atualiza o título da aba com o título da página"""
        index = self.sites_tabs.indexOf(browser)
        if index >= 0:
            self.sites_tabs.setTabText(index, title[:30] + "..." if len(title) > 30 else title)

    def close_site_tab(self, index):
        """Fecha a aba especificada com confirmação"""
        # Verifica se a aba é a "Relação de Sites"
        if index == 0:
            QMessageBox.information(
                self,
                "Aviso",
                "A aba 'Relação de Sites' não pode ser fechada."
            )
            return

        # Obter o widget antes de remover
        widget = self.sites_tabs.widget(index)
        
        # Exibir mensagem de confirmação
        reply = QMessageBox.question(
            self,
            "Remover site?",
            "Tem certeza que deseja remover este site?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Remover a aba
            self.sites_tabs.removeTab(index)
            
            # Limpar o widget
            if widget:
                widget.deleteLater()

    def show_add_dialog(self):
        """Mostra diálogo para adicionar site"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Adicionar Site")
        dialog.setFixedWidth(300)
        
        layout = QVBoxLayout()
        
        name_input = QLineEdit()
        url_input = QLineEdit()
        name_input.setPlaceholderText("Nome do site")
        url_input.setPlaceholderText("URL do site")
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | 
            QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        
        layout.addWidget(QLabel("Nome:"))
        layout.addWidget(name_input)
        layout.addWidget(QLabel("URL:"))
        layout.addWidget(url_input)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            name = name_input.text().strip()
            url = url_input.text().strip()
            
            if name and url:
                item = QListWidgetItem(f"{name} - {url}")
                item.setData(Qt.ItemDataRole.UserRole, {"name": name, "url": url})
                self.sites_list.addItem(item)
                self.save_sites()

    def remove_site(self):
        """Remove o site selecionado"""
        current_item = self.sites_list.currentItem()
        if current_item:
            reply = QMessageBox.question(
                self, 
                'Confirmar Remoção',
                'Deseja remover este site?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.sites_list.takeItem(self.sites_list.row(current_item))
                self.save_sites()

    def preview_site(self, item):
        """Abre o site em uma nova aba no visualizador"""
        if item:
            site_data = item.data(Qt.ItemDataRole.UserRole)
            url = site_data['url']
            name = site_data['name']
            
            if not url.startswith(('http://', 'https://')):
                url = 'https://' + url
            
            web_view = QWebEngineView()
            web_view.setUrl(QUrl(url))
            
            self.preview_tabs.addTab(web_view, name)
            self.preview_tabs.setCurrentIndex(self.preview_tabs.count() - 1)

    def close_preview_tab(self, index):
        """Fecha uma aba do visualizador"""
        self.preview_tabs.removeTab(index)

    def save_sites(self):
        """Salva a lista de sites em JSON"""
        sites = []
        for i in range(self.sites_list.count()):
            item = self.sites_list.item(i)
            sites.append(item.data(Qt.ItemDataRole.UserRole))
        
        with open('sites_list.json', 'w', encoding='utf-8') as f:
            json.dump(sites, f, ensure_ascii=False, indent=4)

    def load_sites(self):
        """Carrega sites do arquivo JSON"""
        if os.path.exists('sites_list.json'):
            with open('sites_list.json', 'r', encoding='utf-8') as f:
                sites = json.load(f)
                for site in sites:
                    item = QListWidgetItem(f"{site['name']} - {site['url']}")
                    item.setData(Qt.ItemDataRole.UserRole, site)
                    self.sites_list.addItem(item)

    def load_sites(self):
        """Carrega sites do arquivo JSON"""
        try:
            with open("sites.json", "r", encoding="utf-8") as f:
                self.sites = json.load(f)
                self.update_sites_list()
        except FileNotFoundError:
            self.sites = []

    def save_sites(self):
        """Salva sites no arquivo JSON"""
        with open("sites.json", "w", encoding="utf-8") as f:
            json.dump(self.sites, f, indent=4, ensure_ascii=False)
        self.update_sites_list()

    def update_sites_list(self):
        """Atualiza a lista de sites na interface"""
        self.sites_list.clear()
        for site in self.sites:
            self.sites_list.addItem(site["nome"])

    def show_add_site_dialog(self):
        """Mostra diálogo para adicionar site"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Adicionar Site")
        dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout()
        
        # Campos de entrada
        name_input = QLineEdit()
        name_input.setPlaceholderText("Nome do site")
        url_input = QLineEdit()
        url_input.setPlaceholderText("URL do site")
        
        layout.addWidget(QLabel("Nome:"))
        layout.addWidget(name_input)
        layout.addWidget(QLabel("URL:"))
        layout.addWidget(url_input)
        
        # Botões
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | 
            QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            nome = name_input.text().strip()
            url = url_input.text().strip()
            if nome and url:
                self.sites.append({"nome": nome, "url": url})
                self.save_sites()

    def show_remove_site_dialog(self):
        """Mostra diálogo para remover site"""
        current_item = self.sites_list.currentItem()
        if current_item:
            reply = QMessageBox.question(
                self,
                "Confirmar Remoção",
                "Deseja remover este site?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                index = self.sites_list.currentRow()
                self.sites.pop(index)
                self.save_sites()
                self.site_details.hide()

    def show_edit_site_dialog(self):
        """Mostra diálogo para editar site"""
        current_item = self.sites_list.currentItem()
        if not current_item:
            return
        
        try:
            # Carregar sites do arquivo JSON
            with open('sites.json', 'r', encoding='utf-8') as f:
                sites = json.load(f)
            
            # Encontrar o site pelo nome atual
            old_name = current_item.text()
            site = next((s for s in sites if s['nome'] == old_name), None)
            
            if not site:
                return
            
            dialog = QDialog(self)
            dialog.setWindowTitle("Editar Site")
            dialog.setMinimumWidth(400)
            
            layout = QVBoxLayout()
            
            # Campos de entrada
            name_input = QLineEdit(site["nome"])
            url_input = QLineEdit(site["url"])
            
            layout.addWidget(QLabel("Nome:"))
            layout.addWidget(name_input)
            layout.addWidget(QLabel("URL:"))
            layout.addWidget(url_input)
            
            # Botões
            buttons = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Save | 
                QDialogButtonBox.StandardButton.Cancel
            )
            layout.addWidget(buttons)
            
            dialog.setLayout(layout)
            
            def save_changes():
                new_name = name_input.text().strip()
                new_url = url_input.text().strip()
                
                if not new_name or not new_url:
                    QMessageBox.warning(dialog, "Erro", "Nome e URL são obrigatórios!")
                    return
                
                try:
                    # Atualizar site no array
                    site['nome'] = new_name
                    site['url'] = new_url
                    
                    # Salvar no arquivo
                    with open('sites.json', 'w', encoding='utf-8') as f:
                        json.dump(sites, f, ensure_ascii=False, indent=4)
                    
                    # Atualizar item na lista
                    current_item.setText(new_name)
                    
                    # Atualizar título da aba se estiver aberta
                    for i in range(self.sites_tabs.count()):
                        tab_text = self.sites_tabs.tabText(i)
                        if tab_text == old_name:
                            print(f"Atualizando aba de '{old_name}' para '{new_name}'")
                            self.sites_tabs.setTabText(i, new_name)
                            
                            # Forçar atualização visual
                            self.sites_tabs.update()
                            QApplication.processEvents()
                        
                    QMessageBox.information(dialog, "Sucesso", "Site atualizado com sucesso!")
                    dialog.accept()
                    
                except Exception as e:
                    QMessageBox.warning(dialog, "Erro", f"Erro ao salvar alterações: {str(e)}")
            
            buttons.accepted.connect(save_changes)
            buttons.rejected.connect(dialog.reject)
            
            dialog.exec()
            
        except Exception as e:
            print(f"Erro ao editar site: {str(e)}")
            QMessageBox.warning(self, "Erro", f"Erro ao editar site: {str(e)}")

    def export_sites(self):
        """Exporta sites para arquivo JSON"""
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar Sites",
            "",
            "Arquivos JSON (*.json)"
        )
        
        if file_name:
            try:
                with open(file_name, "w", encoding="utf-8") as f:
                    json.dump(self.sites, f, indent=4, ensure_ascii=False)
                QMessageBox.information(self, "Sucesso", "Sites exportados com sucesso!")
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao exportar sites: {str(e)}")

    def import_sites(self):
        """Importa sites de arquivo JSON"""
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Importar Sites",
            "",
            "Arquivos JSON (*.json)"
        )
        
        if file_name:
            try:
                with open(file_name, "r", encoding="utf-8") as f:
                    imported_sites = json.load(f)
                    self.sites.extend(imported_sites)
                    self.save_sites()
                QMessageBox.information(self, "Sucesso", "Sites importados com sucesso!")
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao importar sites: {str(e)}")

    def on_site_selected(self, item):
        """Callback quando um site é selecionado na lista"""
        try:
            if item:
                site_info = item.text()
                
                # Criar container para a aba
                container = QWidget()
                layout = QVBoxLayout(container)
                layout.setContentsMargins(0, 0, 0, 0)  # Remove todas as margens
                layout.setSpacing(0)  # Remove espaçamento entre elementos
                
                # Criar barra de navegação
                nav_bar = QHBoxLayout()
                nav_bar.setContentsMargins(1, 0, 1, 1)  # Reduz ainda mais as margens, especialmente no topo
                nav_bar.setSpacing(8)  # Reduz espaçamento entre botões
                
                # Botões de navegação
                back_button = QPushButton("Voltar")
                back_button.setMaximumWidth(60)
                forward_button = QPushButton("Avançar")
                forward_button.setMaximumWidth(60)
                refresh_button = QPushButton("Atualizar")
                refresh_button.setMaximumWidth(60)
                
                # Campo de endereço
                address_bar = QLineEdit()
                address_bar.setPlaceholderText("Digite o endereço da web")
                
                # Adicionar widgets à barra de navegação
                nav_bar.addWidget(back_button)
                nav_bar.addWidget(forward_button)
                nav_bar.addWidget(refresh_button)
                nav_bar.addWidget(address_bar)
                
                # Criar widget para conter a barra de navegação
                nav_widget = QWidget()
                nav_widget.setLayout(nav_bar)
                nav_widget.setContentsMargins(0, 0, 0, 0)  # Remove todas as margens do widget de navegação
                nav_widget.setFixedHeight(50)  # Define altura fixa para a barra de navegação
                
                # Carregar sites do arquivo JSON
                with open('sites.json', 'r', encoding='utf-8') as f:
                    sites = json.load(f)
                
                # Encontrar o site selecionado
                site = next((s for s in sites if s['nome'] == site_info), None)
                
                if site:
                    url = site['url']
                    if not url.startswith(('http://', 'https://')):
                        url = 'https://' + url
                    
                    # Criar visualizador web
                    web_view = QWebEngineView()
                    web_view.setUrl(QUrl(url))
                    web_view.setContentsMargins(0, 0, 0, 0)
                    
                    # Conectar botões e campo de endereço
                    back_button.clicked.connect(web_view.back)
                    forward_button.clicked.connect(web_view.forward)
                    refresh_button.clicked.connect(web_view.reload)
                    address_bar.returnPressed.connect(
                        lambda: self.load_url(web_view, address_bar.text())
                    )
                    
                    # Atualizar o campo de endereço quando a URL mudar
                    web_view.urlChanged.connect(lambda qurl: address_bar.setText(qurl.toString()))
                    
                    # Adicionar widgets ao layout
                    layout.addWidget(nav_widget)
                    layout.addWidget(web_view)
                    
                    # Verificar se já existe uma aba com este site
                    for i in range(self.sites_tabs.count()):
                        if self.sites_tabs.tabText(i) == site_info:
                            self.sites_tabs.setCurrentIndex(i)
                            return
                    
                    # Adicionar nova aba
                    index = self.sites_tabs.addTab(container, site_info)
                    self.sites_tabs.setCurrentIndex(index)
                    
        except Exception as e:
            print(f"Erro ao carregar site: {str(e)}")
            QMessageBox.warning(
                self,
                "Erro",
                f"Erro ao carregar site: {str(e)}"
            )

    def open_in_browser(self):
        """Abre o site selecionado no navegador padrão"""
        current_item = self.sites_list.currentItem()
        if current_item:
            index = self.sites_list.currentRow()
            site = self.sites[index]
            webbrowser.open(site['url'])

    def close_site_tab(self, index):
        """Fecha uma aba do visualizador"""
        self.sites_tabs.removeTab(index)

    def mark_selected_item(self):
        """Marca o item selecionado na lista de mensagens."""
        item = self.messages_list.currentItem()
        if item:
            item.setBackground(QBrush(QColor("yellow")))
            self.save_marked_items()

    def unmark_selected_item(self):
        """Desmarca o item selecionado na lista de mensagens."""
        item = self.messages_list.currentItem()
        if item:
            item.setBackground(QBrush(QColor("white")))
            self.save_marked_items()

    def save_marked_items(self):
        """Salva o estado das marcações em um arquivo JSON."""
        if self.active_button:
            button_name = self.active_button.text()
            marked_items = []
            for index in range(self.messages_list.count()):
                item = self.messages_list.item(index)
                if item.background().color() == QColor("yellow"):
                    marked_items.append(item.text())
            
            try:
                with open('marked_items.json', 'r', encoding='utf-8') as f:
                    all_marked_items = json.load(f)
                    if not isinstance(all_marked_items, dict):
                        all_marked_items = {}
            except (FileNotFoundError, json.JSONDecodeError):
                all_marked_items = {}

            all_marked_items[button_name] = marked_items

            with open('marked_items.json', 'w', encoding='utf-8') as f:
                json.dump(all_marked_items, f, ensure_ascii=False, indent=4)

    def load_marked_items(self):
        """Carrega o estado das marcações de um arquivo JSON."""
        if self.active_button:
            button_name = self.active_button.text()
            try:
                with open('marked_items.json', 'r', encoding='utf-8') as f:
                    all_marked_items = json.load(f)
                    if not isinstance(all_marked_items, dict):
                        all_marked_items = {}
                    marked_items = all_marked_items.get(button_name, [])
                    for index in range(self.messages_list.count()):
                        item = self.messages_list.item(index)
                        if item.text() in marked_items:
                            item.setBackground(QBrush(QColor("yellow")))
            except FileNotFoundError:
                pass  # Se o arquivo não existir, não faz nada
            except Exception as e:
                print(f"Erro ao carregar marcações: {str(e)}")

    def add_new_tab(self, url="https://www.google.com", label="Nova Aba"):
        """Adiciona uma nova aba com um visualizador de site."""
        site_viewer = QWidget()
        layout = QVBoxLayout(site_viewer)
        
        # Barra de navegação
        nav_bar = QHBoxLayout()
        
        back_button = QPushButton("<")
        forward_button = QPushButton(">")
        address_bar = QLineEdit()
        address_bar.setText(url)
        
        nav_bar.addWidget(back_button)
        nav_bar.addWidget(forward_button)
        nav_bar.addWidget(address_bar)
        
        # Visualizador de site
        web_view = QWebEngineView()
        web_view.setUrl(QUrl(url))
        
        # Conectar botões e campo de endereço
        back_button.clicked.connect(lambda: web_view.back())
        forward_button.clicked.connect(lambda: web_view.forward())
        address_bar.returnPressed.connect(lambda: web_view.setUrl(QUrl(address_bar.text())))
        web_view.urlChanged.connect(lambda qurl: address_bar.setText(qurl.toString()))
        
        layout.addLayout(nav_bar)
        layout.addWidget(web_view)
        
        index = self.browser_tabs.insertTab(self.browser_tabs.count() - 1, site_viewer, label)
        self.browser_tabs.setCurrentIndex(index)

    def add_plus_tab(self):
        """Adiciona uma aba especial que atua como botão de adição."""
        plus_tab = QWidget()
        index = self.browser_tabs.addTab(plus_tab, "+")
        self.browser_tabs.tabBar().setTabButton(index, QTabBar.ButtonPosition.LeftSide, None)
        self.browser_tabs.tabBar().setTabButton(index, QTabBar.ButtonPosition.RightSide, None)
        self.browser_tabs.tabBarClicked.connect(self.handle_plus_tab_click)

    def handle_plus_tab_click(self, index):
        """Manipula o clique na aba de adição."""
        if index == self.browser_tabs.count() - 1:  # Se a aba clicada for a última (aba de adição)
            self.add_new_tab("https://www.google.com", "Nova Aba")  # Adiciona uma nova aba com o site padrão

    def close_tab(self, index):
        """Fecha a aba no índice especificado."""
        if self.browser_tabs.count() > 2:  # Manter pelo menos uma aba de site e a aba de adição
            self.browser_tabs.removeTab(index)

    def add_new_browser_tab(self, url="https://www.google.com", label="Nova Aba"):
        """Adiciona uma nova aba de navegador com um QWebEngineView."""
        # Criar layout para a barra de navegação
        nav_bar = QHBoxLayout()

        # Botões de navegação
        back_button = QPushButton("Voltar")
        forward_button = QPushButton("Avançar")
        refresh_button = QPushButton("Atualizar")

        # Campo de endereço
        address_bar = QLineEdit()
        address_bar.setPlaceholderText("Digite o endereço da web")

        # Criar o visualizador de site
        web_view = QWebEngineView()
        web_view.setUrl(QUrl(url))

        # Conectar botões e campo de endereço
        back_button.clicked.connect(lambda: web_view.back())
        forward_button.clicked.connect(lambda: web_view.forward())
        refresh_button.clicked.connect(lambda: web_view.reload())
        address_bar.returnPressed.connect(lambda: self.load_url(web_view, address_bar.text()))

        # Atualizar o campo de endereço quando a URL mudar
        web_view.urlChanged.connect(lambda qurl: self.update_tab_title(web_view, qurl.toString()))

        # Adicionar widgets à barra de navegação
        nav_bar.addWidget(back_button)
        nav_bar.addWidget(forward_button)
        nav_bar.addWidget(refresh_button)
        nav_bar.addWidget(address_bar)

        # Criar um layout vertical para a aba
        tab_layout = QVBoxLayout()
        tab_layout.addLayout(nav_bar)
        tab_layout.addWidget(web_view)

        # Criar um widget para a aba e definir o layout
        tab_widget = QWidget()
        tab_widget.setLayout(tab_layout)

        # Adicionar a nova aba ao QTabWidget
        index = self.browser_tabs.insertTab(self.browser_tabs.count() - 1, tab_widget, label)
        self.browser_tabs.setCurrentIndex(index)

    def load_url(self, web_view, url):
        """Carrega a URL no QWebEngineView, ajustando o formato se necessário."""
        if not url.startswith(('http://', 'https://')):
            url = 'http://' + url
        web_view.setUrl(QUrl(url))

    def handle_tab_click(self, index):
        """Manipula o clique na aba de adição."""
        if index == self.browser_tabs.count() - 1:  # Se a aba clicada for a última (aba de adição)
            self.add_new_browser_tab()  # Adiciona uma nova aba com o site padrão

    def close_browser_tab(self, index):
        """Fecha a aba do navegador no índice especificado."""
        if self.browser_tabs.count() > 2:  # Manter pelo menos uma aba aberta e a aba de adição
            self.browser_tabs.removeTab(index)

    def save_tabs(self):
        """Salva as URLs das abas abertas em um arquivo JSON."""
        urls = []
        for i in range(self.browser_tabs.count() - 1):  # Ignorar a aba de adição
            tab_widget = self.browser_tabs.widget(i)
            web_view = tab_widget.findChild(QWebEngineView)
            if web_view:
                urls.append(web_view.url().toString())

        with open('tabs.json', 'w', encoding='utf-8') as f:
            json.dump(urls, f, ensure_ascii=False, indent=4)

    def load_tabs(self):
        """Carrega as URLs das abas salvas de um arquivo JSON."""
        try:
            with open('tabs.json', 'r', encoding='utf-8') as f:
                urls = json.load(f)
                for url in urls:
                    self.add_new_browser_tab(url)
        except FileNotFoundError:
            # Se o arquivo não existir, adicionar uma aba padrão
            self.add_new_browser_tab("https://www.google.com", "Google")

        # Adicionar a aba de adição
        self.add_plus_tab()

    def closeEvent(self, event):
        """Salva o estado das abas ao fechar o programa."""
        self.save_tabs()
        super().closeEvent(event)

    def update_tab_title(self, web_view, url):
        """Atualiza o título da aba com a URL da página."""
        index = self.browser_tabs.indexOf(web_view.parentWidget())
        if index != -1:
            # Trunca a URL se for muito longa
            truncated_url = url if len(url) <= 30 else url[:27] + "..."
            self.browser_tabs.setTabText(index, truncated_url)

class PDFViewer(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Barra de controles
        controls = QHBoxLayout()
        
        # Campo de pesquisa
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Pesquisar...")
        self.search_input.setFixedWidth(200)
        
        # Botões de pesquisa
        self.search_prev = QPushButton("↑")
        self.search_next = QPushButton("↓")
        self.search_prev.setFixedSize(30, 30)
        self.search_next.setFixedSize(30, 30)
        
        # Adicionar widgets à barra de controles (removidos os botões de navegação)
        controls.addStretch()
        controls.addWidget(self.search_input)
        controls.addWidget(self.search_prev)
        controls.addWidget(self.search_next)
        controls.addStretch()
        
        # Área de visualização do PDF usando WebEngine
        self.web_view = QWebEngineView()
        settings = self.web_view.page().settings()
        settings.setAttribute(QWebEngineSettings.WebAttribute.PluginsEnabled, True)
        settings.setAttribute(QWebEngineSettings.WebAttribute.PdfViewerEnabled, True)
        
        # Conectar sinais (apenas pesquisa)
        self.search_input.returnPressed.connect(self.search_text)
        self.search_prev.clicked.connect(self.search_previous)
        self.search_next.clicked.connect(self.search_next_text)
        
        layout.addLayout(controls)
        layout.addWidget(self.web_view)

    def load_pdf(self, pdf_path):
        """Carrega um arquivo PDF no visualizador"""
        if pdf_path and os.path.exists(pdf_path):
            url = QUrl.fromLocalFile(pdf_path)
            self.web_view.setUrl(url)
        else:
            QMessageBox.warning(self, "Erro", "Arquivo PDF não encontrado!")

    def search_text(self):
        text = self.search_input.text()
        if text:
            self.web_view.page().findText(text)

    def search_previous(self):
        text = self.search_input.text()
        if text:
            self.web_view.page().findText(text, QWebEnginePage.FindFlag.FindBackward)

    def search_next_text(self):
        text = self.search_input.text()
        if text:
            self.web_view.page().findText(text)

class CustomTabWidget(QTabWidget):
    def __init__(self):
        super().__init__()
        self.setTabsClosable(True)
        self.setMovable(True)
        
        # Estilo atualizado
        self.setStyleSheet("""
            QTabWidget::pane {
                border: none;
            }
            QTabBar::tab {
                padding: 5px 25px 5px 10px;
                margin-right: 2px;
                border: 1px solid #ccc;
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                background-color: #f0f0f0;
            }
            QTabBar::tab:selected {
                background-color: white;
            }
            QTabBar::close-button {
                image: url(close.png);
                subcontrol-position: right;
                subcontrol-origin: padding;
                margin-right: 4px;
                position: absolute;
                right: 4px;
            }
            QTabBar::close-button:hover {
                background-color: #ff4444;
                border-radius: 2px;
            }
        """)
        
        # Criar o botão de adicionar aba
        self.add_tab_button = QToolButton(self)
        self.add_tab_button.setText("+")
        self.add_tab_button.setFixedSize(25, 25)
        self.add_tab_button.clicked.connect(self.add_new_tab)
        self.add_tab_button.setStyleSheet("""
            QToolButton {
                border: none;
                background-color: transparent;
                font-size: 16px;
                padding: 2px;
                margin: 0;
            }
            QToolButton:hover {
                background-color: #e0e0e0;
                border-radius: 2px;
            }
        """)
        
        # Conectar sinais
        self.tabCloseRequested.connect(self.close_tab)
        self.tabBar().tabMoved.connect(self.update_add_button_position)
        
        # Criar 3 abas iniciais
        for i in range(3):
            self.add_new_tab()
            
        # Posicionar o botão de adicionar
        self.update_add_button_position()

    def update_add_button_position(self):
        """Atualiza a posição do botão + para ficar após a última aba"""
        tab_bar = self.tabBar()
        if tab_bar.count() > 0:
            last_tab_rect = tab_bar.tabRect(tab_bar.count() - 1)
            x = last_tab_rect.right() + 5
            y = last_tab_rect.top() + (last_tab_rect.height() - self.add_tab_button.height()) // 2
            self.add_tab_button.move(x, y)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_add_button_position()

    def tabInserted(self, index):
        super().tabInserted(index)
        self.update_add_button_position()

    def tabRemoved(self, index):
        super().tabRemoved(index)
        self.update_add_button_position()

    def add_new_tab(self):
        """Adiciona uma nova aba com um QWebEngineView"""
        web_view = QWebEngineView()
        web_view.setUrl(QUrl("https://www.google.com"))
        
        # Configurar a página web com configurações simplificadas
        settings = web_view.settings()
        settings.setAttribute(QWebEngineSettings.WebAttribute.PluginsEnabled, False)
        settings.setAttribute(QWebEngineSettings.WebAttribute.JavascriptCanOpenWindows, False)
        
        index = self.addTab(web_view, f"Nova aba")
        self.setCurrentIndex(index)
        
        # Atualizar o título quando a página carregar
        web_view.titleChanged.connect(lambda title, view=web_view: 
            self.setTabText(self.indexOf(view), title[:20] + "..." if len(title) > 20 else title))

    def close_tab(self, index):
        """Solicita confirmação e fecha a aba"""
        if self.count() > 1:  # Manter pelo menos uma aba aberta
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Question)
            msg.setText("Deseja fechar esta aba?")
            msg.setWindowTitle("Confirmar fechamento")
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            
            if msg.exec_() == QMessageBox.Yes:
                self.removeTab(index)
        else:
            QMessageBox.warning(self, "Aviso", "Não é possível fechar a última aba.")

def remove_selected_items():
    try:
        items_to_remove = []
        current_segment = None
        
        # Percorrer todos os layouts no right_scroll_layout
        i = 0
        while i < right_scroll_layout.count():
            item = right_scroll_layout.itemAt(i)
            if isinstance(item, QHBoxLayout):
                checkbox = item.itemAt(0).widget()
                label = item.itemAt(1).widget()
                
                if isinstance(checkbox, QCheckBox) and isinstance(label, QLabel):
                    text = label.text()
                    
                    # Verificar se é um segmento
                    if text in self.segmentos:
                        current_segment = text
                        if checkbox.isChecked():
                            items_to_remove.append(('segment', text))
                    # Se não é segmento e tem segmento atual, é um botão
                    elif current_segment and checkbox.isChecked():
                        items_to_remove.append(('button', (current_segment, text)))
                i += 1
        
        if items_to_remove:
            msg = QMessageBox(dialog)
            msg.setIcon(QMessageBox.Icon.Question)
            msg.setWindowTitle("Confirmar Remoção")
            msg.setText("Deseja remover os itens selecionados?")
            msg.setStandardButtons(
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if msg.exec() == QMessageBox.StandardButton.Yes:
                # Remover itens
                for item_type, item_data in items_to_remove:
                    if item_type == 'segment':
                        segment_name = item_data
                        if segment_name in self.segmentos:
                            # Remover frases dos botões
                            for button in self.segmentos[segment_name]:
                                if button in self.frases:
                                    del self.frases[button]
                            # Remover segmento
                            del self.segmentos[segment_name]
                    else:
                        segment_name, button_name = item_data
                        if segment_name in self.segmentos and button_name in self.segmentos[segment_name]:
                            self.segmentos[segment_name].remove(button_name)
                            if button_name in self.frases:
                                del self.frases[button_name]
                
                # Remover segmentos vazios
                for segment in list(self.segmentos.keys()):
                    if not self.segmentos[segment]:
                        del self.segmentos[segment]
                
                # Salvar e atualizar
                self.save_phrases_to_file()
                update_right_panel()
                
                QMessageBox.information(
                    dialog,
                    "Sucesso",
                    "Itens removidos com sucesso!"
                )
    except Exception as e:
        QMessageBox.warning(
            dialog,
            "Erro",
            f"Erro ao remover itens: {str(e)}"
        )

if __name__ == '__main__':
    try:
        # Força a liberação de memória antes de iniciar
        import gc
        gc.collect()
        
        # Configura o uso de memória alta prioridade
        app = QApplication(sys.argv)
        app.setStyle('Fusion')
        
        # Configura prioridade de processamento
        import psutil
        process = psutil.Process()
        process.nice(psutil.HIGH_PRIORITY_CLASS)
        
        # Aloca memória para a aplicação
        app.processEvents()
        
        # Monitora uso de memória
        def monitor_memory():
            memory_info = process.memory_info()
            print(f"Uso de memória: {memory_info.rss / 1024 / 1024:.2f} MB")
        
        # Timer para monitorar memória
        timer = QTimer()
        timer.timeout.connect(monitor_memory)
        timer.start(10000)  # Monitora a cada 10 segundos
        
        # Inicia a janela principal
        window = MainWindow()
        window.showMaximized()  # Janela maximizada ao iniciar
        
        # Inicia o loop de eventos - Alterado de exec_ para exec
        sys.exit(app.exec())
        
    except Exception as e:
        print(f"Erro na inicialização: {str(e)}")
        sys.exit(1)
    finally:
        # Limpa a memória ao fechar
        gc.collect()


 
