Tela inicial-Layout PyQT 4.1.py

import os
import sys
import json
import shutil
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import fitz
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtWebEngineWidgets import QWebEngineView, QWebEngineSettings
from PyQt5.QtWidgets import (QDialog, QTableWidget, QTableWidgetItem, 
                            QDialogButtonBox, QVBoxLayout, QMessageBox)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtCore import QTimer

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Visualização de Pagamentos")
        self.setGeometry(100, 100, 1200, 800)

        # Inicializar variáveis
        self.active_button = None
        self.active_frases = None
        self.frases = {}
        
        # Dicionário de PDFs organizado (movido para antes do init_ui)
        self.pdf_buttons = {
            "Alteração dos dados do cartão de crédito": "",
            "Card de Propostas Pendentes": "",
            "Chamado - Canal de apoio ao Corretor": "",
            "Chat suporte ao corretor": "",
            "Condições Gerais": "",
            "Critérios de bônus": "",
            "Endosso Online": "",
            "Enquadramento de veículo em tempo de Campanha": "",
            "Manual consulta restrições": "",
            "Manual de extensões de perímetro": "",
            "Manual de PGIT": "",
            "Manual de Regras de Bônus": "",
            "Manual do Corretor": "",
            "Pagamento PGIT": "",
            "Passo a passo Livelo - Corretor": "",
            "Passo a passo Livelo - Matriz": "",
            "Retificação de propostas": "",
            "Retirada de boleto": "",
            "Segunda Chance": "",
            "Tela retificadora do corretor": "",
            "Vistoria Prévia": ""
        }
        
        self.segmentos = {
            "Pagamentos em Andamento": [
                "Boleto",
                "Débito em conta",
                "Cartão de crédito",
                "Endosso - Débito",
                "Endosso - Cartão",
                "CONTRATO - IND"
            ],
            "Segunda Chance": [
                "Boleto - Segunda Chance",
                "Débito em Conta - Segunda Chance",
                "Carto de crédito - Segunda Chance",
                "Endosso Débito - Segunda Chance",
                "Endosso Boleto - Segunda Chance"
            ],
            "Pagamentos e Livelo": [
                "Pagamentos - Pagamentos e Livelo",
                "Livelo - Pagamentos e Livelo"
            ],
            "Classe de Bônus": [
                "Classe de bônus",
                "Canc. Falta de Pagamento",
                "Ren. Sem Sinistro igual ou maior - 335 dias",
                "Ren. Sem Sinistro igual ou menor 335 dias",
                "Leitura de Bônus"
            ],
            "Vistoria": [
                "Posto Fixo",
                "Análise especial"
            ],
            "Rastreador": [
                "Ordem não gerada",
                "Ordem gerada",
                "Ordem cancelada"
            ],
            "Apólice": [
                "Apólice (Geral)",
                "Emissão de boleto",
                "Cartão de crédito"
            ]
        }

        # Carregar dados
        self.load_phrases_from_file()
        self.load_pdf_links()
        
        # Inicializar interface
        self.init_ui()

    def init_ui(self):
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Layout principal com tabs
        self.tab_widget = QTabWidget()
        
        # Primeira aba (Chat)
        primeira_aba = QWidget()
        primeira_layout = QHBoxLayout()
        
        # Container esquerdo
        container_esquerdo = QWidget()
        layout_esquerdo = QVBoxLayout()
        
        # Adiciona o visualizador de site
        site_view = QWebEngineView()
        site_view.setUrl(QUrl("https://www.google.com.br"))
        layout_esquerdo.addWidget(site_view)
        
        # Lista de mensagens com filtro
        filter_layout = self.create_filter_layout()
        layout_esquerdo.addLayout(filter_layout)
        
        self.messages_list = QListWidget()
        self.messages_list.itemClicked.connect(self.copy_to_clipboard)
        self.messages_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.messages_list.customContextMenuRequested.connect(self.show_context_menu)
        layout_esquerdo.addWidget(self.messages_list)
        
        # Novo frame para botões de alerta
        alert_frame = QFrame()
        alert_frame.setFrameStyle(QFrame.Box | QFrame.Plain)
        alert_frame.setLineWidth(1)
        alert_buttons_layout = QHBoxLayout(alert_frame)
        alert_buttons_layout.setContentsMargins(10, 5, 10, 5)
        
        # Estilo para os botões
        button_style = """
            QPushButton {
                padding: 5px 15px;
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 3px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #e0e0e0;
            }
        """
        
        # Criar botões
        alerta8_btn = QPushButton("Alerta 8")
        alerta9_btn = QPushButton("Alerta 9")
        ppe_btn = QPushButton("PPE")
        sinistro_btn = QPushButton("Sinistro")
        
        # Adicionar checkbox para controle de imagens, anexos e ausências
        self.remove_images_cb = QCheckBox("Retirar imagens")
        self.remove_attachments_cb = QCheckBox("Remover anexo")
        self.evandro_absent_cb = QCheckBox("Evandro - férias-ausente")
        self.aline_absent_cb = QCheckBox("Aline - férias-ausente")
        
        # Estilo para os checkboxes
        checkbox_style = """
            QCheckBox {
                margin-left: 10px;
                padding: 5px;
            }
        """
        self.remove_images_cb.setStyleSheet(checkbox_style)
        self.remove_attachments_cb.setStyleSheet(checkbox_style)
        self.evandro_absent_cb.setStyleSheet(checkbox_style)
        self.aline_absent_cb.setStyleSheet(checkbox_style)
        
        # Aplicar estilo aos botões
        for btn in [alerta8_btn, alerta9_btn, ppe_btn, sinistro_btn]:
            btn.setStyleSheet(button_style)
        
        # Adicionar botões e checkboxes ao layout
        alert_buttons_layout.addWidget(alerta8_btn)
        alert_buttons_layout.addWidget(alerta9_btn)
        alert_buttons_layout.addWidget(ppe_btn)
        alert_buttons_layout.addWidget(sinistro_btn)
        alert_buttons_layout.addWidget(self.remove_images_cb)
        alert_buttons_layout.addWidget(self.remove_attachments_cb)
        alert_buttons_layout.addWidget(self.evandro_absent_cb)
        alert_buttons_layout.addWidget(self.aline_absent_cb)
        alert_buttons_layout.addStretch()
        
        # Conectar sinais dos botões
        alerta8_btn.clicked.connect(lambda: self.open_outlook_email("Alerta 8"))
        alerta9_btn.clicked.connect(lambda: self.open_outlook_email("Alerta 9"))
        ppe_btn.clicked.connect(lambda: self.open_outlook_email("PPE"))
        sinistro_btn.clicked.connect(lambda: self.open_outlook_email("Sinistro"))
        
        # Adicionar frame ao layout
        layout_esquerdo.addWidget(alert_frame)
        
        container_esquerdo.setLayout(layout_esquerdo)
        primeira_layout.addWidget(container_esquerdo, 1)
        
        # Área de botões
        primeira_layout.addWidget(self.create_buttons_area(), 0)
        
        primeira_aba.setLayout(primeira_layout)
        self.tab_widget.addTab(primeira_aba, "Chat")
        
        # Segunda aba (Manuais)
        segunda_aba = QWidget()
        segunda_layout = QHBoxLayout()
        
        # Área do visualizador de PDF (lado esquerdo)
        pdf_viewer_group = QGroupBox("Visualizador")
        pdf_viewer_layout = QVBoxLayout()
        
        # Substituir o QLabel pelo novo PDFViewer
        self.pdf_viewer = PDFViewer()
        pdf_viewer_layout.addWidget(self.pdf_viewer)
        
        pdf_viewer_group.setLayout(pdf_viewer_layout)
        segunda_layout.addWidget(pdf_viewer_group, 4)
        
        # Área de botões (lado direito)
        buttons_group = QGroupBox("Manuais e Documentos")
        buttons_layout = QVBoxLayout()
        
        # Checkbox para modo de edição
        self.edit_mode_cb = QCheckBox("Modo de Edição")
        self.edit_mode_cb.setStyleSheet("""
            QCheckBox {
                padding: 5px;
                font-weight: bold;
            }
        """)
        buttons_layout.addWidget(self.edit_mode_cb)
        
        # Scroll area para os botões
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        # Estilo para os botões
        button_style = """
            QPushButton {
                text-align: left;
                padding: 10px;
                border: 1px solid #ccc;
                border-radius: 5px;
                background-color: #f8f9fa;
            }
            QPushButton:hover {
                background-color: #e9ecef;
            }
        """
        
        # Adicionar botões para cada PDF
        for nome_pdf in self.pdf_buttons.keys():
            btn = QPushButton(nome_pdf)
            btn.setFixedHeight(35)
            btn.setMinimumWidth(50)
            btn.setStyleSheet(button_style)
            btn.clicked.connect(lambda checked, name=nome_pdf: self.handle_pdf_button_click(name))
            scroll_layout.addWidget(btn)
        
        scroll_layout.addStretch()
        scroll_widget.setLayout(scroll_layout)
        
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        buttons_layout.addWidget(scroll)
        
        # Botão de remover
        remove_pdf_button = QPushButton("Remover PDFs")
        remove_pdf_button.clicked.connect(self.show_remove_dialog)
        buttons_layout.addWidget(remove_pdf_button)
        
        buttons_group.setLayout(buttons_layout)
        segunda_layout.addWidget(buttons_group, 1)
        
        segunda_aba.setLayout(segunda_layout)
        self.tab_widget.addTab(segunda_aba, "Manuais")
        
        # Terceira aba (Visualizador TXT)
        terceira_aba = QWidget()
        terceira_layout = QHBoxLayout()
        
        # Área do visualizador de texto (lado esquerdo)
        text_viewer_group = QGroupBox("Visualizador de Texto")
        text_viewer_layout = QVBoxLayout()
        
        # Configuração do QTextEdit com suporte a múltipla seleção
        self.text_viewer = QTextEdit()
        self.text_viewer.setReadOnly(True)
        self.text_viewer.setAcceptRichText(False)
        self.text_viewer.setLineWrapMode(QTextEdit.NoWrap)
        self.text_viewer.setUndoRedoEnabled(False)
        self.selected_texts = []
        
        # Configurar o comportamento de seleção múltipla
        self.text_viewer.mousePressEvent = self.custom_mouse_press_event
        self.text_viewer.mouseReleaseEvent = self.custom_mouse_release_event
        
        text_viewer_layout.addWidget(self.text_viewer)
        text_viewer_group.setLayout(text_viewer_layout)
        terceira_layout.addWidget(text_viewer_group, 2)
        
        # Área de botões (lado direito)
        buttons_group = QGroupBox("Controles")
        buttons_layout = QVBoxLayout()
        
        load_button = QPushButton("Carregar Arquivo")
        load_button.clicked.connect(self.load_txt_file)
        
        copy_button = QPushButton("Copiar para Chat")
        copy_button.clicked.connect(self.show_copy_dialog)
        
        buttons_layout.addWidget(load_button)
        buttons_layout.addWidget(copy_button)
        buttons_layout.addStretch()
        
        buttons_group.setLayout(buttons_layout)
        terceira_layout.addWidget(buttons_group, 1)
        
        terceira_aba.setLayout(terceira_layout)
        self.tab_widget.addTab(terceira_aba, "Visualizador TXT")
        
        # Quarta aba (Navegador)
        quarta_aba = QWidget()
        quarta_layout = QVBoxLayout()
        
        # Container para as abas e botão de adicionar
        tab_container = QWidget()
        tab_container_layout = QHBoxLayout()
        tab_container_layout.setSpacing(0)
        tab_container_layout.setContentsMargins(0, 0, 0, 0)
        
        # Widget de abas personalizadas
        self.browser_tabs = CustomTabWidget()
        tab_container_layout.addWidget(self.browser_tabs)
        
        tab_container_layout.addStretch()
        
        tab_container.setLayout(tab_container_layout)
        quarta_layout.addWidget(tab_container)
        
        quarta_aba.setLayout(quarta_layout)
        self.tab_widget.addTab(quarta_aba, "Navegador")
        
        # Layout principal
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.tab_widget)
        central_widget.setLayout(main_layout)
        
        # Barra de status
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

    def set_active_button(self, button, frases):
        if self.active_button:
            self.active_button.setStyleSheet("")
        self.active_button = button
        self.active_frases = frases
        self.active_button.setStyleSheet("background-color: lightblue;")
        self.mostrar_frases()

    def copy_to_clipboard(self, item):
        """Copia a frase para a área de transferência"""
        clipboard = QApplication.clipboard()
        clipboard.setText(item.text())

    def load_phrases_from_file(self):
        try:
            with open('frases.json', 'r', encoding='utf-8') as file:
                data = json.load(file)
                self.frases = data.get('frases', {})
                loaded_segmentos = data.get('segmentos', {})
                if loaded_segmentos:
                    self.segmentos.update(loaded_segmentos)
        except FileNotFoundError:
            print("Arquivo de frases não encontrado. Criando novo arquivo.")
            self.save_phrases_to_file()
        except Exception as e:
            print(f"Erro ao carregar frases: {str(e)}")

    def save_phrases_to_file(self):
        try:
            data = {
                'frases': self.frases,
                'segmentos': self.segmentos
            }
            with open('frases.json', 'w', encoding='utf-8') as file:
                json.dump(data, file, ensure_ascii=False, indent=4)
            return True
        except Exception as e:
            print(f"Erro ao salvar frases: {str(e)}")
            return False

    def create_buttons_area(self):
        """Cria a área de botões"""
        container = QWidget()
        main_layout = QVBoxLayout(container)
        
        # Ajustando margens do layout principal
        main_layout.setContentsMargins(25, 15, 25, 15)  # Margens para centralizar conteúdo
        main_layout.setSpacing(8)
        
        # Botões de controle
        control_layout = QHBoxLayout()
        add_button = QPushButton("+")
        remove_button = QPushButton("-")
        edit_button = QPushButton("Edição")
        add_button.setFixedSize(30, 30)
        remove_button.setFixedSize(30, 30)
        edit_button.setFixedSize(60, 30)
        add_button.clicked.connect(self.add_button_clicked)
        remove_button.clicked.connect(self.remove_button_clicked)
        edit_button.clicked.connect(self.edit_button_clicked)
        control_layout.addWidget(add_button)
        control_layout.addWidget(remove_button)
        control_layout.addWidget(edit_button)
        control_layout.addStretch()
        main_layout.addLayout(control_layout)
        
        # Criar botões por segmento
        for segmento, botoes in self.segmentos.items():
            label = QLabel(segmento)
            label.setStyleSheet("font-weight: bold; margin-top: 5px;")
            main_layout.addWidget(label)
            
            section_layout = QVBoxLayout()
            section_layout.setSpacing(8)
            section_layout.setContentsMargins(5, 0, 5, 0)
            
            for botao in botoes:
                if botao not in self.frases:
                    self.frases[botao] = []
                
                btn = QPushButton(botao)
                btn.setFixedSize(340, 50)
                btn.clicked.connect(lambda checked, b=btn, f=self.frases[botao]: 
                                  self.set_active_button(b, f))
                section_layout.addWidget(btn)
            
            main_layout.addLayout(section_layout)
            main_layout.addSpacing(5)
        
        # Área de rolagem com borda visível
        scroll = QScrollArea()
        scroll.setWidget(container)
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("""
            QScrollArea {
                border: 1px solid #999;
                border-radius: 0px;
                background-color: white;
                margin: 5px;
            }
        """)
        
        # Widget para envolver o ScrollArea e garantir a margem
        wrapper = QWidget()
        wrapper_layout = QVBoxLayout(wrapper)
        wrapper_layout.setContentsMargins(5, 5, 5, 5)
        wrapper_layout.addWidget(scroll)
        
        return wrapper

    def add_button_clicked(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Adicionar Novo Botão")
        dialog.setMinimumWidth(600)
        
        # Layout principal horizontal
        main_layout = QHBoxLayout()
        
        # Layout esquerdo para Novo Segmento
        left_layout = QVBoxLayout()
        
        # Container para "Novo Segmento"
        new_segment_container = QVBoxLayout()
        
        # Layout horizontal para checkbox e label
        checkbox_layout = QHBoxLayout()
        new_segment_checkbox = QCheckBox()
        new_segment_label = QLabel("Novo Segmento")
        new_segment_label.setStyleSheet("font-weight: bold;")
        checkbox_layout.addWidget(new_segment_checkbox)
        checkbox_layout.addWidget(new_segment_label)
        checkbox_layout.addStretch()
        new_segment_container.addLayout(checkbox_layout)
        
        # Campo para nome do segmento (inicialmente oculto)
        segment_input = QLineEdit()
        segment_input.setPlaceholderText("Digite o nome do novo segmento")
        segment_input.hide()
        new_segment_container.addWidget(segment_input)
        
        # Campo para nome do botão
        button_label = QLabel("Nome do Botão:")
        button_input = QLineEdit()
        button_input.setPlaceholderText("Digite o nome do botão")
        new_segment_container.addWidget(button_label)
        new_segment_container.addWidget(button_input)
        
        left_layout.addLayout(new_segment_container)
        left_layout.addStretch()
        
        # Frame separador vertical
        separator = QFrame()
        separator.setFrameShape(QFrame.VLine)
        separator.setFrameShadow(QFrame.Sunken)
        
        # Layout direito para segmentos existentes
        right_layout = QVBoxLayout()
        
        # Lista de segmentos existentes com checkboxes
        for segmento, botoes in self.segmentos.items():
            segment_container = QHBoxLayout()
            segment_checkbox = QCheckBox()
            segment_container.addWidget(segment_checkbox)
            segment_label = QLabel(f"{segmento}:")
            segment_label.setStyleSheet("font-weight: bold;")
            segment_container.addWidget(segment_label)
            segment_container.addStretch()
            right_layout.addLayout(segment_container)
            
            for botao in botoes:
                button_container = QHBoxLayout()
                button_container.addSpacing(20)
                button_label = QLabel(botao)
                button_container.addWidget(button_label)
                button_container.addStretch()
                right_layout.addLayout(button_container)
        
        # Botões de confirmação no layout principal
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        
        # Layout final
        final_layout = QVBoxLayout()
        main_layout.addLayout(left_layout, 1)
        main_layout.addWidget(separator)
        main_layout.addLayout(right_layout, 2)
        final_layout.addLayout(main_layout)
        final_layout.addWidget(button_box)
        
        dialog.setLayout(final_layout)
        
        # Função para mostrar/ocultar campo de segmento
        def toggle_segment_input():
            segment_input.setVisible(new_segment_checkbox.isChecked())
            dialog.adjustSize()

        new_segment_checkbox.stateChanged.connect(toggle_segment_input)
        
        if dialog.exec_() == QDialog.Accepted:
            button_name = button_input.text().strip()
            if button_name:
                if new_segment_checkbox.isChecked():
                    new_segment = segment_input.text().strip()
                    if new_segment:
                        if new_segment not in self.segmentos:
                            self.segmentos[new_segment] = []
                        self.segmentos[new_segment].append(button_name)
                else:
                    for i in range(right_layout.count()):
                        item = right_layout.itemAt(i)
                        if isinstance(item, QHBoxLayout):
                            checkbox = item.itemAt(0).widget()
                            if isinstance(checkbox, QCheckBox) and checkbox.isChecked():
                                label = item.itemAt(1).widget()
                                segment_name = label.text().replace(':', '')
                                self.segmentos[segment_name].append(button_name)
                                break
                
                self.frases[button_name] = []
                self.save_phrases_to_file()
                self.init_ui()

    def remove_button_clicked(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Remover")
        layout = QVBoxLayout()

        # Dicionário para armazenar os checkboxes dos segmentos
        segment_checkboxes = {}

        # Lista de segmentos e botões com checkboxes
        for segmento, botoes in self.segmentos.items():
            # Container horizontal para checkbox e label do segmento
            segment_container = QHBoxLayout()
            
            # Checkbox para o segmento
            segment_checkbox = QCheckBox()
            segment_container.addWidget(segment_checkbox)
            segment_checkboxes[segmento] = segment_checkbox
            
            # Label do segmento
            segment_label = QLabel(f"{segmento}:")
            segment_label.setStyleSheet("font-weight: bold;")
            segment_container.addWidget(segment_label)
            
            # Adiciona espaço flexível
            segment_container.addStretch()
            
            layout.addLayout(segment_container)
            
            # Checkboxes para os botões (com indentaço)
            for botao in botoes:
                button_container = QHBoxLayout()
                button_container.addSpacing(20)  # Indentação
                
                # Checkbox do botão
                button_checkbox = QCheckBox()
                button_container.addWidget(button_checkbox)
                
                # Label do botão
                button_label = QLabel(botao)
                button_container.addWidget(button_label)
                
                # Adiciona espaço flexível
                button_container.addStretch()
                
                layout.addLayout(button_container)

        # Botões de confirmação
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        dialog.setLayout(layout)

        if dialog.exec_() == QDialog.Accepted:
            # Verificar segmentos marcados
            segments_to_remove = []
            for segmento, checkbox in segment_checkboxes.items():
                if checkbox.isChecked():
                    segments_to_remove.append(segmento)
            
            # Se houver segmentos marcados, pedir confirmação
            if segments_to_remove:
                segments_list = "\n".join(segments_to_remove)
                reply = QMessageBox.question(
                    dialog,
                    'Confirmar Exclusão',
                    f'Deseja realmente excluir os seguintes segmentos e todos os seus botões?\n\n{segments_list}',
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    # Remove os segmentos confirmados
                    for segmento in segments_to_remove:
                        for botao in self.segmentos[segmento]:
                            if botao in self.frases:
                                del self.frases[botao]
                        del self.segmentos[segmento]

            # Processar os botões marcados
            for segmento, botoes in self.segmentos.copy().items():
                button_containers = [layout.itemAt(i).layout() for i in range(layout.count())
                                   if isinstance(layout.itemAt(i), QHBoxLayout)]
                
                for container in button_containers:
                    if container and container.count() >= 2:
                        checkbox = container.itemAt(1).widget()
                        if isinstance(checkbox, QCheckBox) and checkbox.isChecked():
                            button_name = container.itemAt(2).widget().text()
                            if button_name in botoes:
                                botoes.remove(button_name)
                                if button_name in self.frases:
                                    del self.frases[button_name]

            # Atualiza a interface e salva as alterações
            self.save_phrases_to_file()
            self.init_ui()

    def update_buttons_area(self):
        """Atualiza a área de botões após modificações"""
        primeira_aba = self.tab_widget.widget(0)
        if primeira_aba and primeira_aba.layout():
            layout = primeira_aba.layout()
            
            # Remover área de botões antiga
            if layout.count() > 1:
                old_buttons = layout.takeAt(1)
                if old_buttons and old_buttons.widget():
                    old_buttons.widget().deleteLater()
            
            # Adicionar nova área de botões
            layout.addWidget(self.create_buttons_area(), 0)

    def add_pdf(self, pdf_name=None):
        """Adiciona ou atualiza um PDF"""
        if pdf_name is None:
            pdf_name, ok = QInputDialog.getItem(
                self,
                "Selecionar Manual",
                "Escolha o manual para adicionar:",
                [k for k, v in self.pdf_buttons.items() if not v],
                0,
                False
            )
            if not ok:
                return

        file_name, _ = QFileDialog.getOpenFileName(
            self,
            f"Selecionar PDF para {pdf_name}",
            "",
            "PDF Files (*.pdf)"
        )
        
        if file_name:
            try:
                # Criar diretório pdfs se não existir
                if not os.path.exists('pdfs'):
                    os.makedirs('pdfs')
                
                # Copiar arquivo para o diretório pdfs
                new_name = f"{pdf_name}.pdf"
                new_path = os.path.join('pdfs', new_name)
                shutil.copy2(file_name, new_path)
                
                # Atualizar dicionário com caminho absoluto
                self.pdf_buttons[pdf_name] = os.path.abspath(new_path)
                self.save_pdf_links()
                
                # Tentar abrir o PDF imediatamente
                self.pdf_viewer.load_pdf(self.pdf_buttons[pdf_name])
                
                self.status_bar.showMessage(f"PDF '{pdf_name}' adicionado com sucesso!", 2000)
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao adicionar PDF: {str(e)}")

    def remove_pdf(self):
        """Remove o vínculo com um PDF"""
        pdf_name, ok = QInputDialog.getItem(
            self,
            "Remover PDF",
            "Escolha o manual para remover:",
            [k for k, v in self.pdf_buttons.items() if v],  # Mostrar apenas os com arquivo
            0,
            False
        )
        if ok and pdf_name:
            try:
                pdf_path = self.pdf_buttons[pdf_name]
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                
                self.pdf_buttons[pdf_name] = ""
                self.save_pdf_links()
                
                self.status_bar.showMessage(f"PDF '{pdf_name}' removido com sucesso!", 2000)
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao remover PDF: {str(e)}")

    def save_pdf_links(self):
        """Salva os links dos PDFs em um arquivo"""
        try:
            with open('pdf_links.json', 'w', encoding='utf-8') as f:
                json.dump(self.pdf_buttons, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Erro ao salvar links dos PDFs: {str(e)}")

    def load_pdf_links(self):
        """Carrega os links dos PDFs do arquivo"""
        try:
            if os.path.exists('pdf_links.json'):
                with open('pdf_links.json', 'r', encoding='utf-8') as f:
                    saved_links = json.load(f)
                    self.pdf_buttons.update(saved_links)
        except Exception as e:
            print(f"Erro ao carregar links dos PDFs: {str(e)}")

    def open_pdf(self, pdf_name):
        """Abre o PDF selecionado"""
        pdf_path = self.pdf_buttons.get(pdf_name)
        print(f"Tentando abrir PDF: {pdf_name}")
        print(f"Caminho do arquivo: {pdf_path}")
        
        if not pdf_path:
            QMessageBox.warning(
                self,
                "PDF não encontrado",
                f"O PDF '{pdf_name}' ainda não foi vinculado. Por favor, adicione o arquivo."
            )
            if self.edit_mode_cb.isChecked():
                self.add_pdf(pdf_name)
        else:
            try:
                if os.path.exists(pdf_path):
                    print(f"Arquivo existe em: {pdf_path}")
                    # Carregar PDF no visualizador
                    self.pdf_viewer.load_pdf(pdf_path)
                    self.status_bar.showMessage(f"PDF carregado: {pdf_name}", 2000)
                else:
                    print(f"Arquivo não encontrado em: {pdf_path}")
                    QMessageBox.warning(self, "Erro", "Arquivo PDF não encontrado!")
            except Exception as e:
                print(f"Erro ao abrir PDF: {str(e)}")
                QMessageBox.warning(self, "Erro", f"Erro ao abrir PDF: {str(e)}")

    def update_button_combo(self, segment):
        """Atualiza o combo de botões baseado no segmento selecionado"""
        self.button_combo.clear()
        if segment in self.segmentos:
            self.button_combo.addItems(self.segmentos[segment])
        self.update_edit_list()

    def update_edit_list(self):
        """Atualiza a lista de frases do botão selecionado"""
        self.edit_list.clear()
        button = self.button_combo.currentText()
        if button in self.frases:
            self.edit_list.addItems(self.frases[button])

    def add_phrase(self):
        """Adiciona uma nova frase ao botão selecionado"""
        button = self.button_combo.currentText()
        if button:
            text, ok = QInputDialog.getText(self, 'Adicionar Frase', 'Digite a nova frase:')
            if ok and text:
                if button not in self.frases:
                    self.frases[button] = []
                self.frases[button].append(text)
                self.save_phrases_to_file()
                self.update_edit_list()
                self.status_bar.showMessage("Frase adicionada com sucesso!", 2000)

    def edit_phrase(self):
        """Edita a frase selecionada"""
        current_item = self.edit_list.currentItem()
        if current_item:
            button = self.button_combo.currentText()
            old_text = current_item.text()
            text, ok = QInputDialog.getText(self, 'Editar Frase', 'Edite a frase:', text=old_text)
            if ok and text:
                idx = self.frases[button].index(old_text)
                self.frases[button][idx] = text
                self.save_phrases_to_file()
                self.update_edit_list()
                self.status_bar.showMessage("Frase editada com sucesso!", 2000)

    def remove_phrase(self):
        """Remove a frase selecionada"""
        current_item = self.edit_list.currentItem()
        if current_item:
            button = self.button_combo.currentText()
            text = current_item.text()
            reply = QMessageBox.question(self, 'Confirmar Remoção', 
                                       'Tem certeza que deseja remover esta frase?',
                                       QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.frases[button].remove(text)
                self.save_phrases_to_file()
                self.update_edit_list()
                self.status_bar.showMessage("Frase removida com sucesso!", 2000)

    def open_text_file(self):
        """Abre um arquivo de texto"""
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Abrir Arquivo", "", "Arquivos de Texto (*.txt);;Todos os Arquivos (*)"
        )
        if file_name:
            try:
                with open(file_name, 'r', encoding='utf-8') as file:
                    self.edit_text.setPlainText(file.read())
                self.status_bar.showMessage(f"Arquivo aberto: {file_name}", 2000)
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao abrir arquivo: {str(e)}")

    def save_text_file(self):
        """Salva o arquivo de texto"""
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Salvar Arquivo", "", "Arquivos de Texto (*.txt);;Todos os Arquivos (*)"
        )
        if file_name:
            try:
                with open(file_name, 'w', encoding='utf-8') as file:
                    file.write(self.edit_text.toPlainText())
                self.status_bar.showMessage(f"Arquivo salvo: {file_name}", 2000)
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao salvar arquivo: {str(e)}")

    def handle_pdf_button_click(self, pdf_name):
        """Gerencia o clique nos botões de PDF"""
        if self.edit_mode_cb.isChecked():
            self.add_pdf(pdf_name)
        else:
            self.open_pdf(pdf_name)

    def show_remove_dialog(self):
        """Mostra diálogo para remover PDFs"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Remover PDFs")
        dialog.setMinimumWidth(400)
        layout = QVBoxLayout()

        # Criar lista de checkboxes
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout()
        
        checkboxes = {}
        for pdf_name, pdf_path in self.pdf_buttons.items():
            cb = QCheckBox(pdf_name)
            # Adicionar indicador visual se tem PDF vinculado
            if pdf_path:
                cb.setText(f"{pdf_name} (PDF vinculado)")
                cb.setStyleSheet("QCheckBox { color: green; }")
            else:
                cb.setText(f"{pdf_name} (Sem PDF)")
                cb.setStyleSheet("QCheckBox { color: gray; }")
            
            checkboxes[pdf_name] = cb
            scroll_layout.addWidget(cb)
        
        scroll_widget.setLayout(scroll_layout)
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        
        # Adicionar botões de selecionar/desselecionar todos
        select_layout = QHBoxLayout()
        select_all_btn = QPushButton("Selecionar Todos")
        deselect_all_btn = QPushButton("Desselecionar Todos")
        
        def select_all():
            for cb in checkboxes.values():
                cb.setChecked(True)
        
        def deselect_all():
            for cb in checkboxes.values():
                cb.setChecked(False)
        
        select_all_btn.clicked.connect(select_all)
        deselect_all_btn.clicked.connect(deselect_all)
        
        select_layout.addWidget(select_all_btn)
        select_layout.addWidget(deselect_all_btn)
        
        layout.addLayout(select_layout)
        layout.addWidget(scroll)
        
        # Botões OK/Cancelar
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        
        if dialog.exec_() == QDialog.Accepted:
            # Remover PDFs selecionados
            removed = False
            for pdf_name, cb in checkboxes.items():
                if cb.isChecked():
                    try:
                        pdf_path = self.pdf_buttons[pdf_name]
                        if pdf_path and os.path.exists(pdf_path):
                            os.remove(pdf_path)
                        self.pdf_buttons[pdf_name] = ""
                        removed = True
                    except Exception as e:
                        QMessageBox.warning(self, "Erro", f"Erro ao remover {pdf_name}: {str(e)}")
            
            if removed:
                self.save_pdf_links()
                self.status_bar.showMessage("PDFs selecionados foram removidos com sucesso!", 2000)
                # Atualizar a visualização
                self.pdf_viewer.doc = None
                self.pdf_viewer.pdf_label.clear()
                self.pdf_viewer.pdf_label.setText("Selecione um manual para visualizar")

    def create_filter_layout(self):
        filter_layout = QHBoxLayout()
        
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Digite para filtrar as mensagens...")
        self.filter_input.textChanged.connect(self.filter_messages)
        
        add_button = QPushButton("+")
        add_button.setFixedSize(30, 30)
        add_button.clicked.connect(self.add_frase)
        
        remove_button = QPushButton("-")
        remove_button.setFixedSize(30, 30)
        remove_button.clicked.connect(self.remove_frase)
        
        view_button = QPushButton("Ver Planilha")
        view_button.setFixedSize(100, 30)
        view_button.clicked.connect(self.view_planilha)
        
        export_button = QPushButton("Exportar para Excel")
        export_button.setFixedSize(150, 30)
        export_button.clicked.connect(self.export_to_excel)
        
        # Botão Abrir com tamanho reduzido
        open_folder_button = QPushButton("Abrir")
        open_folder_button.setFixedSize(50, 30)  # Reduzido de 60 para 50 pixels
        open_folder_button.clicked.connect(self.open_export_folder)
        
        filter_layout.addWidget(self.filter_input)
        filter_layout.addWidget(add_button)
        filter_layout.addWidget(remove_button)
        filter_layout.addWidget(view_button)
        filter_layout.addWidget(export_button)
        filter_layout.addWidget(open_folder_button)
        
        return filter_layout

    def filter_messages(self, text):
        for index in range(self.messages_list.count()):
            item = self.messages_list.item(index)
            item.setHidden(text.lower() not in item.text().lower())

    def view_planilha(self):
        """Abre uma janela de diálogo para visualizar as frases em formato de tabela"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Planilha de Dados")
        dialog_layout = QVBoxLayout()

        if self.active_frases:
            # Criar tabela
            table = QTableWidget()
            table.setRowCount(len(self.active_frases))
            table.setColumnCount(1)
            table.setHorizontalHeaderLabels(["Frases"])
            table.horizontalHeader().setStretchLastSection(True)

            # Preencher tabela com as frases
            for row, frase in enumerate(sorted(self.active_frases)):
                table.setItem(row, 0, QTableWidgetItem(frase))

            # Configurar a tabela para permitir edição
            table.cellChanged.connect(lambda row, col: self.update_phrase(row, col, table))
            
            dialog_layout.addWidget(table)
            
            # Adicionar botão OK
            button_box = QDialogButtonBox(QDialogButtonBox.Ok)
            button_box.accepted.connect(dialog.accept)
            dialog_layout.addWidget(button_box)

            dialog.setLayout(dialog_layout)
            dialog.resize(600, 400)  # Tamanho inicial da janela
            dialog.exec_()
        else:
            QMessageBox.information(self, "Aviso", "Nenhuma frase disponível para visualização.")

    def update_phrase(self, row, col, table):
        """Atualiza a frase quando editada na tabela"""
        if self.active_frases and row < len(self.active_frases):
            item = table.item(row, col)
            if item:
                new_text = item.text()
                self.active_frases[row] = new_text
                self.active_frases = sorted(self.active_frases)
                self.frases[self.active_button.text()] = self.active_frases
                self.mostrar_frases()
                self.save_phrases_to_file()

    def export_to_excel(self):
        """Exporta as frases para um arquivo Excel"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Salvar Planilha", 
            "", 
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if file_path:
            try:
                # Criar novo workbook
                workbook = Workbook()
                workbook.remove(workbook.active)  # Remove a planilha padrão

                # Ordenar os nomes dos botões alfabeticamente
                botoes_ordenados = sorted(self.frases.keys())

                for button_name in botoes_ordenados:
                    button_frases = self.frases[button_name]
                    if not button_frases:
                        continue  # Pula botões sem frases
                    
                    # Criar nova sheet com o nome do botão
                    sheet = workbook.create_sheet(title=button_name)
                    
                    # Adicionar título (nome do botão) na primeira linha
                    header_cell = sheet.cell(row=1, column=1, value=button_name.upper())
                    header_cell.font = Font(name='Calibri', size=8, bold=True)
                    header_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Encontrar a maior frase para ajustar a largura da coluna
                    max_length = len(button_name)  # Inicializa com o tamanho do título
                    for frase in button_frases:
                        max_length = max(max_length, len(frase))
                    
                    # Ajustar largura da coluna baseado na maior frase
                    # Fator de ajuste para a fonte Calibri tamanho 8
                    sheet.column_dimensions['A'].width = max_length * 0.9
                    
                    # Adicionar frases
                    for idx, frase in enumerate(sorted(button_frases), start=2):
                        cell = sheet.cell(row=idx, column=1, value=frase)
                        cell.font = Font(name='Calibri', size=8)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Ajustar altura das linhas para centralização vertical
                    for row in sheet.rows:
                        sheet.row_dimensions[row[0].row].height = 15

                workbook.save(file_path)
                QMessageBox.information(self, "Sucesso", "Planilha exportada com sucesso!")
                
                # Salva o caminho do arquivo exportado
                settings = QSettings('MyApp', 'ExcelExport')
                settings.setValue('last_export_path', file_path)
                
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao exportar planilha:\n{str(e)}")

    def add_frase(self):
        text = self.filter_input.text()
        if text and self.active_frases is not None:
            self.active_frases.append(text)
            self.active_frases = sorted(self.active_frases)
            self.frases[self.active_button.text()] = self.active_frases
            self.mostrar_frases()
            self.filter_input.clear()
            self.save_phrases_to_file()

    def remove_frase(self):
        selected_items = self.messages_list.selectedItems()
        if selected_items:
            for item in selected_items:
                self.active_frases.remove(item.text())
            self.active_frases = sorted(self.active_frases)
            self.frases[self.active_button.text()] = self.active_frases
            self.mostrar_frases()
            self.save_phrases_to_file()

    def mostrar_frases(self):
        self.messages_list.clear()
        for frase in sorted(self.active_frases):
            item = QListWidgetItem(frase)
            self.messages_list.addItem(item)

    def show_context_menu(self, position):
        menu = QMenu(self)
        remove_action = menu.addAction("Remover Frase")
        action = menu.exec_(self.messages_list.mapToGlobal(position))
        if action == remove_action:
            self.remove_frase()

    def load_txt_file(self):
        """Carrega um arquivo TXT para visualização"""
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Abrir Arquivo TXT",
            "",
            "Arquivos de Texto (*.txt);;Todos os Arquivos (*)"
        )
        if file_name:
            try:
                with open(file_name, 'r', encoding='utf-8') as file:
                    self.text_viewer.setText(file.read())
                self.status_bar.showMessage(f"Arquivo carregado: {file_name}", 2000)
            except Exception as e:
                QMessageBox.warning(self, "Erro", f"Erro ao carregar arquivo: {str(e)}")

    def show_copy_dialog(self):
        """Mostra diálogo para selecionar destino da cópia"""
        if not self.selected_texts:
            cursor = self.text_viewer.textCursor()
            if cursor.hasSelection():
                self.selected_texts = [cursor.selectedText()]
        
        if not self.selected_texts:
            QMessageBox.warning(self, "Aviso", "Por favor, selecione um texto para copiar.")
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Copiar para Chat")
        dialog.setMinimumWidth(400)
        layout = QVBoxLayout()
        
        # Label de instrução com contador de frases
        instruction = QLabel(f"Selecione o destino para {len(self.selected_texts)} frase(s):")
        layout.addWidget(instruction)
        
        # Preview das frases selecionadas
        preview_group = QGroupBox("Texto selecionado")
        preview_layout = QVBoxLayout()
        preview_text = QTextEdit()
        preview_text.setPlainText('\n'.join(self.selected_texts))
        preview_text.setReadOnly(True)
        preview_text.setMaximumHeight(100)
        preview_layout.addWidget(preview_text)
        preview_group.setLayout(preview_layout)
        layout.addWidget(preview_group)
        
        # Scroll area para os checkboxes
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout()
        
        checkboxes = {}
        # Criar checkboxes para cada botão em cada segmento
        for segmento, botoes in self.segmentos.items():
            # Label do segmento
            segment_label = QLabel(segmento)
            segment_label.setStyleSheet("font-weight: bold;")
            scroll_layout.addWidget(segment_label)
            
            # Checkboxes dos botões
            for botao in botoes:
                cb = QCheckBox(botao)
                checkboxes[botao] = cb
                scroll_layout.addWidget(cb)
            
            scroll_layout.addSpacing(10)
        
        scroll_widget.setLayout(scroll_layout)
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        layout.addWidget(scroll)
        
        # Botões de seleção rápida
        select_buttons = QHBoxLayout()
        select_all = QPushButton("Selecionar Todos")
        deselect_all = QPushButton("Desselecionar Todos")
        
        def select_all_boxes():
            for cb in checkboxes.values():
                cb.setChecked(True)
            
        def deselect_all_boxes():
            for cb in checkboxes.values():
                cb.setChecked(False)
        
        select_all.clicked.connect(select_all_boxes)
        deselect_all.clicked.connect(deselect_all_boxes)
        
        select_buttons.addWidget(select_all)
        select_buttons.addWidget(deselect_all)
        layout.addLayout(select_buttons)
        
        # Botões OK/Cancelar
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        
        if dialog.exec_() == QDialog.Accepted:
            # Processar seleções
            copied_to = []
            frases_duplicadas = []
            
            for botao, cb in checkboxes.items():
                if cb.isChecked():
                    if botao not in self.frases:
                        self.frases[botao] = []
                    
                    # Adicionar cada frase ao botão, verificando duplicatas
                    for frase in self.selected_texts:
                        if frase:
                            # Verifica se a frase já existe no botão
                            if frase not in self.frases[botao]:
                                self.frases[botao].append(frase)
                            else:
                                if botao not in frases_duplicadas:
                                    frases_duplicadas.append(botao)
                
                    # Ordenar as frases
                    self.frases[botao] = sorted(self.frases[botao])
                    copied_to.append(botao)
                    
                    # Se este é o botão ativo, atualizar a visualização
                    if self.active_button and self.active_button.text() == botao:
                        self.mostrar_frases()
            
            # Limpar seleções após copiar
            self.selected_texts.clear()
            self.text_viewer.setExtraSelections([])
            
            if copied_to:
                # Salvar alterações
                if self.save_phrases_to_file():
                    # Preparar mensagem de status
                    msg_parts = []
                    if len(copied_to) > len(frases_duplicadas):
                        msg_parts.append(f"Frases copiadas para: {', '.join(set(copied_to) - set(frases_duplicadas))}")
                    if frases_duplicadas:
                        msg_parts.append(f"Frases já existentes em: {', '.join(frases_duplicadas)}")
                    
                    msg = " | ".join(msg_parts)
                    self.status_bar.showMessage(msg, 5000)
                    
                    # Atualizar a visualização se houver um botão ativo
                    if self.active_button:
                        self.active_frases = self.frases[self.active_button.text()]
                        self.mostrar_frases()
                else:
                    QMessageBox.warning(
                        self, 
                        "Erro", 
                        "Erro ao salvar as frases. Verifique as permissões do arquivo."
                    )
            else:
                QMessageBox.warning(
                    self, 
                    "Aviso", 
                    "Nenhum destino selecionado para copiar as frases."
                )

    def eventFilter(self, obj, event):
        if obj is self.text_viewer and event.type() == QEvent.KeyPress:
            if event.key() == Qt.Key_Control:
                # Habilita seleção múltipla quando CTRL é pressionado
                self.text_viewer.setTextInteractionFlags(
                    Qt.TextSelectableByMouse | Qt.TextSelectableByKeyboard
                )
                return True
        elif obj is self.text_viewer and event.type() == QEvent.KeyRelease:
            if event.key() == Qt.Key_Control:
                # Desabilita seleção múltipla quando CTRL é liberado
                self.text_viewer.setTextInteractionFlags(Qt.TextSelectableByMouse)
                return True
        return super().eventFilter(obj, event)

    def setup_text_viewer(self):
        """Configura o visualizador de texto com suporte a múltiplas seleções"""
        self.text_viewer = QTextEdit()
        self.text_viewer.setReadOnly(True)
        self.text_viewer.installEventFilter(self)
        self.text_viewer.setTextInteractionFlags(Qt.TextSelectableByMouse)

    def custom_mouse_press_event(self, event):
        """Manipula o evento de pressionar o mouse no visualizador de texto"""
        if event.button() == Qt.LeftButton:
            cursor = self.text_viewer.cursorForPosition(event.pos())
            if event.modifiers() & Qt.ControlModifier:
                # Se CTRL está pressionado, seleciona a linha inteira
                cursor.movePosition(QTextCursor.StartOfLine)
                cursor.movePosition(QTextCursor.EndOfLine, QTextCursor.KeepAnchor)
                selected_text = cursor.selectedText()
                
                if selected_text.strip():
                    # Adiciona a linha selecionada �� lista de seleções
                    if selected_text not in self.selected_texts:
                        self.selected_texts.append(selected_text)
                        
                        # Destaca a linha selecionada
                        extra_selection = QTextEdit.ExtraSelection()
                        extra_selection.format.setBackground(QColor(173, 216, 230))  # Azul claro
                        extra_selection.cursor = cursor
                        
                        current_selections = self.text_viewer.extraSelections()
                        current_selections.append(extra_selection)
                        self.text_viewer.setExtraSelections(current_selections)
            else:
                # Comportamento normal de seleção
                QTextEdit.mousePressEvent(self.text_viewer, event)
                self.selected_texts.clear()
                self.text_viewer.setExtraSelections([])

    def custom_mouse_release_event(self, event):
        """Manipula o evento de soltar o botão do mouse"""
        if event.button() == Qt.LeftButton and not (event.modifiers() & Qt.ControlModifier):
            # Adiciona a seleção normal à lista se não estiver usando CTRL
            cursor = self.text_viewer.textCursor()
            if cursor.hasSelection():
                selected_text = cursor.selectedText()
                if selected_text.strip():
                    self.selected_texts = [selected_text]
            QTextEdit.mouseReleaseEvent(self.text_viewer, event)

    def edit_button_clicked(self):
        """Função para editar os botões existentes"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Editar Botões")
        dialog.setMinimumWidth(400)
        layout = QVBoxLayout()

        # Scroll area para os checkboxes e labels editáveis
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout()

        # Dicionário para armazenar os widgets de edição
        edit_widgets = {}

        # Criar widgets de edição para cada segmento e botão
        for segmento, botoes in self.segmentos.items():
            # Label do segmento
            segment_label = QLabel(segmento)
            segment_label.setStyleSheet("font-weight: bold;")
            scroll_layout.addWidget(segment_label)

            # Criar linha de edição para cada botão
            for botao in botoes:
                button_layout = QHBoxLayout()
                
                # Checkbox para habilitar edição
                checkbox = QCheckBox()
                button_layout.addWidget(checkbox)
                
                # Label editável
                label = QLineEdit(botao)
                label.setReadOnly(True)
                label.setStyleSheet("background-color: #f0f0f0;")
                button_layout.addWidget(label)
                
                # Conectar checkbox com a função de toggle da label
                def toggle_edit(checked, label=label):
                    label.setReadOnly(not checked)
                    label.setStyleSheet("" if checked else "background-color: #f0f0f0;")
                
                checkbox.stateChanged.connect(toggle_edit)
                
                # Armazenar widgets para processamento posterior
                edit_widgets[botao] = {
                    'checkbox': checkbox,
                    'label': label,
                    'segmento': segmento,
                    'original_name': botao
                }
                
                scroll_layout.addLayout(button_layout)
            
            scroll_layout.addSpacing(10)

        scroll_widget.setLayout(scroll_layout)
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        layout.addWidget(scroll)

        # Botões OK/Cancelar
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        dialog.setLayout(layout)

        if dialog.exec_() == QDialog.Accepted:
            try:
                # Processar alterações
                alteracoes = False
                
                # Criar cópias dos dicionários para manipulação
                novos_segmentos = {k: list(v) for k, v in self.segmentos.items()}
                novas_frases = dict(self.frases)
                
                # Mapear alterações de nomes
                nome_alteracoes = {}  # old_name -> new_name
                
                # Processar cada alteração
                for widgets in edit_widgets.values():
                    if widgets['checkbox'].isChecked():
                        old_name = widgets['original_name']
                        new_name = widgets['label'].text().strip()
                        segmento = widgets['segmento']
                        
                        if new_name and new_name != old_name:
                            # Atualizar nome no segmento
                            idx = novos_segmentos[segmento].index(old_name)
                            novos_segmentos[segmento][idx] = new_name
                            
                            # Atualizar frases associadas
                            if old_name in novas_frases:
                                novas_frases[new_name] = novas_frases.pop(old_name)
                            
                            nome_alteracoes[old_name] = new_name
                            alteracoes = True
                
                if alteracoes:
                    # Aplicar as alterações
                    self.segmentos = novos_segmentos
                    self.frases = novas_frases
                    
                    # Salvar no arquivo
                    if self.save_phrases_to_file():
                        # Atualizar os textos dos botões existentes
                        for button in self.findChildren(QPushButton):
                            old_text = button.text()
                            if old_text in nome_alteracoes:
                                button.setText(nome_alteracoes[old_text])
                                
                                # Se este é o botão ativo, atualizar referências
                                if self.active_button and self.active_button == button:
                                    self.active_button = button
                                    self.active_frases = self.frases[nome_alteracoes[old_text]]
                    
                        # Atualizar visualização se necessário
                        if self.active_button:
                            self.mostrar_frases()
                        
                        self.status_bar.showMessage("Botões atualizados com sucesso!", 3000)
                    else:
                        raise Exception("Erro ao salvar as alterações no arquivo")
                    
            except Exception as e:
                QMessageBox.warning(
                    self,
                    "Erro",
                    f"Erro ao atualizar os botões: {str(e)}"
                )

    def open_export_folder(self):
        """Abre o diretório de exportação dos arquivos Excel"""
        try:
            settings = QSettings('MyApp', 'ExcelExport')
            last_export_path = settings.value('last_export_path')
            
            # Define o diretório padrão como 'Downloads' se não houver último caminho
            if last_export_path and os.path.exists(os.path.dirname(last_export_path)):
                target_dir = os.path.dirname(last_export_path)
            else:
                # Tenta usar o diretório Downloads
                if os.name == 'nt':  # Windows
                    target_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
                else:  # Linux/Mac
                    target_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
                
                # Se Downloads não existir, usa o diretório do usuário
                if not os.path.exists(target_dir):
                    target_dir = os.path.expanduser('~')
            
            # Abre o explorador de arquivos no diretório
            if os.name == 'nt':  # Windows
                os.startfile(target_dir)
            else:  # Linux/Mac
                import subprocess
                subprocess.Popen(['xdg-open', target_dir])
            
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao abrir o diretório: {str(e)}")

    def get_greeting(self):
        """Retorna a saudação apropriada baseada no horário atual"""
        from datetime import datetime
        hora_atual = datetime.now().hour
        minuto_atual = datetime.now().minute
        
        # Converte hora e minuto para minutos totais para facilitar a comparação
        tempo_atual = hora_atual * 60 + minuto_atual
        
        # Define os limites dos turnos em minutos
        manha_inicio = 8 * 60 + 30  # 08:30
        manha_fim = 11 * 60 + 59    # 11:59
        tarde_fim = 17 * 60 + 59    # 17:59
        noite_fim = 23 * 60 + 59    # 23:59
        
        if manha_inicio <= tempo_atual <= manha_fim:
            return "Bom dia!"
        elif tempo_atual <= tarde_fim:
            return "Boa tarde!"
        elif tempo_atual <= noite_fim:
            return "Boa noite!"
        else:
            return "Olá!"  # Fora dos horários especificados

    def open_outlook_email(self, search_term):
        try:
            import win32com.client
            outlook = win32com.client.Dispatch("Outlook.Application")
            namespace = outlook.GetNamespace("MAPI")
            
            # Acessar a caixa de saída
            sent_folder = namespace.GetDefaultFolder(5)  # 5 = olFolderSentMail
            
            # Procurar pelo e-mail mais recente com o termo especificado
            items = sent_folder.Items
            items.Sort("[ReceivedTime]", True)  # Ordenar por data, mais recente primeiro
            
            found_email = None
            for item in items:
                if search_term in item.Subject:
                    found_email = item
                    break
            
            if found_email:
                # Criar uma nova mensagem baseada no e-mail encontrado
                new_mail = outlook.CreateItem(0)  # 0 = olMailItem
                new_mail.Subject = found_email.Subject
                
                # Processar o corpo do e-mail baseado no checkbox
                if self.remove_images_cb.isChecked():
                    # Remove tags de imagem do HTML
                    html_body = found_email.HTMLBody
                    import re
                    html_body = re.sub(r'<img[^>]*>', '', html_body)
                    html_body = re.sub(r'<v:imagedata[^>]*>', '', html_body)
                    html_body = re.sub(r'<v:image[^>]*>.*?</v:image>', '', html_body)
                    new_mail.HTMLBody = html_body
                else:
                    new_mail.HTMLBody = found_email.HTMLBody
                
                # Substituir destinatário baseado nos checkboxes de ausência
                original_to = found_email.To
                if search_term == "Alerta 8" and self.evandro_absent_cb.isChecked():
                    original_to = original_to.replace(
                        "evandro.britto@bradescoseguros.com.br",
                        "aline.campos@bradescoseguros.com.br"
                    )
                elif search_term == "Alerta 9" and self.aline_absent_cb.isChecked():
                    original_to = original_to.replace(
                        "aline.campos@bradescoseguros.com.br",
                        "evandro.britto@bradescoseguros.com.br"
                    )
                
                new_mail.To = original_to
                
                # Adicionar saudação apropriada no início do corpo do e-mail
                greeting = self.get_greeting()
                new_mail.HTMLBody = f"{greeting}<br><br>" + new_mail.HTMLBody
                
                # Copiar anexos se o checkbox não estiver marcado
                if not self.remove_attachments_cb.isChecked():
                    for attachment in found_email.Attachments:
                        new_mail.Attachments.Add(attachment.PropertyAccessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F"))
                
                new_mail.Display()  # Mostrar o novo e-mail
            else:
                QMessageBox.warning(self, "Aviso", f"Nenhum e-mail encontrado com o termo '{search_term}'")
                
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao abrir o Outlook: {str(e)}")

class PDFViewer(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_page = 0
        self.zoom_factor = 1.0
        self.doc = None
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Barra de controles
        controls = QHBoxLayout()
        
        # Botões de navegação
        self.prev_btn = QPushButton("←")
        self.next_btn = QPushButton("→")
        self.prev_btn.setFixedSize(40, 30)
        self.next_btn.setFixedSize(40, 30)
        
        # Substituir o QLabel por QLineEdit para permitir digitação
        self.page_input = QLineEdit("0/0")
        self.page_input.setFixedWidth(60)
        self.page_input.setAlignment(Qt.AlignCenter)
        self.page_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 1px solid #ccc;
                border-radius: 3px;
                padding: 2px;
            }
        """)
        self.page_input.returnPressed.connect(self.go_to_page)
        
        # Campo de pesquisa
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Pesquisar...")
        self.search_input.setFixedWidth(200)
        
        # Controles de zoom
        self.zoom_out_btn = QPushButton("-")
        self.zoom_in_btn = QPushButton("+")
        self.zoom_reset_btn = QPushButton("□")
        self.zoom_out_btn.setFixedSize(30, 30)
        self.zoom_in_btn.setFixedSize(30, 30)
        self.zoom_reset_btn.setFixedSize(30, 30)
        
        # Adicionar widgets à barra de controles
        controls.addWidget(self.prev_btn)
        controls.addWidget(self.page_input)
        controls.addWidget(self.next_btn)
        controls.addStretch()
        controls.addWidget(self.search_input)
        controls.addStretch()
        controls.addWidget(self.zoom_out_btn)
        controls.addWidget(self.zoom_reset_btn)
        controls.addWidget(self.zoom_in_btn)
        
        # Área de visualização do PDF
        self.scroll_area = QScrollArea()
        self.pdf_label = QLabel()
        self.pdf_label.setAlignment(Qt.AlignCenter)
        self.scroll_area.setWidget(self.pdf_label)
        self.scroll_area.setWidgetResizable(True)
        
        # Definindo cor de fundo apenas para a área de visualização
        self.scroll_area.setStyleSheet("QScrollArea { background-color: black; }")
        self.pdf_label.setStyleSheet("QLabel { background-color: black; }")
        
        # Conectar sinais
        self.prev_btn.clicked.connect(self.previous_page)
        self.next_btn.clicked.connect(self.next_page)
        self.zoom_in_btn.clicked.connect(lambda: self.zoom(1.2))
        self.zoom_out_btn.clicked.connect(lambda: self.zoom(0.8))
        self.zoom_reset_btn.clicked.connect(self.zoom_reset)
        self.search_input.returnPressed.connect(self.search_text)
        
        # Instalar filtro de eventos para o scroll do mouse
        self.scroll_area.viewport().installEventFilter(self)
        
        layout.addLayout(controls)
        layout.addWidget(self.scroll_area)

    def load_pdf(self, pdf_path):
        print(f"PDFViewer: Tentando carregar {pdf_path}")
        try:
            self.doc = fitz.open(pdf_path)
            print(f"PDFViewer: PDF aberto com {len(self.doc)} páginas")
            self.current_page = 0
            self.zoom_factor = 1.0
            self.update_page_label()
            self.render_page()
        except Exception as e:
            print(f"PDFViewer: Erro ao carregar PDF: {str(e)}")
            QMessageBox.warning(None, "Erro", f"Erro ao carregar PDF: {str(e)}")

    def render_page(self):
        if not self.doc:
            return
            
        try:
            page = self.doc[self.current_page]
            zoom_matrix = fitz.Matrix(1.2 * self.zoom_factor, 1.2 * self.zoom_factor)
            pix = page.get_pixmap(matrix=zoom_matrix)
            
            img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(img)
            
            self.pdf_label.setPixmap(pixmap)
            self.pdf_label.setStyleSheet("background-color: black;")  # Garante fundo preto no label
            self.update_page_label()
        except Exception as e:
            print(f"PDFViewer: Erro ao renderizar página: {str(e)}")

    def update_page_label(self):
        if self.doc:
            self.page_input.setText(f"{self.current_page + 1}/{len(self.doc)}")

    def previous_page(self):
        if self.doc and self.current_page > 0:
            self.current_page -= 1
            self.render_page()

    def next_page(self):
        if self.doc and self.current_page < len(self.doc) - 1:
            self.current_page += 1
            self.render_page()

    def zoom(self, factor):
        self.zoom_factor *= factor
        self.render_page()

    def zoom_reset(self):
        self.zoom_factor = 1.0
        self.render_page()

    def search_text(self):
        if not self.doc:
            return
            
        text = self.search_input.text()
        if not text:
            return
            
        page = self.doc[self.current_page]
        areas = page.search_for(text)
        if areas:
            # Destacar texto encontrado (implementação futura)
            pass

    def eventFilter(self, obj, event):
        if obj is self.scroll_area.viewport():
            if event.type() == QEvent.Wheel:
                # Scroll normal para zoom com Ctrl
                if event.modifiers() & Qt.ControlModifier:
                    if event.angleDelta().y() > 0:
                        self.zoom(1.1)
                    else:
                        self.zoom(0.9)
                    return True
                # Scroll sem Ctrl para navegação de páginas
                else:
                    if event.angleDelta().y() > 0:
                        self.previous_page()
                    else:
                        self.next_page()
                    return True
        return super().eventFilter(obj, event)

    def go_to_page(self):
        if not self.doc:
            return
            
        try:
            text = self.page_input.text()
            if '/' in text:
                page_num = int(text.split('/')[0]) - 1
            else:
                page_num = int(text) - 1
                
            if 0 <= page_num < len(self.doc):
                self.current_page = page_num
                self.render_page()
            else:
                self.update_page_label()  # Restaura o número correto se inválido
        except ValueError:
            self.update_page_label()  # Restaura o número correto se inválido

class CustomTabWidget(QTabWidget):
    def __init__(self):
        super().__init__()
        self.setTabsClosable(True)
        self.setMovable(True)
        
        # Estilo atualizado
        self.setStyleSheet("""
            QTabWidget::pane {
                border: none;
            }
            QTabBar::tab {
                padding: 5px 25px 5px 10px;
                margin-right: 2px;
                border: 1px solid #ccc;
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                background-color: #f0f0f0;
            }
            QTabBar::tab:selected {
                background-color: white;
            }
            QTabBar::close-button {
                image: url(close.png);
                subcontrol-position: right;
                subcontrol-origin: padding;
                margin-right: 4px;
                position: absolute;
                right: 4px;
            }
            QTabBar::close-button:hover {
                background-color: #ff4444;
                border-radius: 2px;
            }
        """)
        
        # Criar o botão de adicionar aba
        self.add_tab_button = QToolButton(self)
        self.add_tab_button.setText("+")
        self.add_tab_button.setFixedSize(25, 25)
        self.add_tab_button.clicked.connect(self.add_new_tab)
        self.add_tab_button.setStyleSheet("""
            QToolButton {
                border: none;
                background-color: transparent;
                font-size: 16px;
                padding: 2px;
                margin: 0;
            }
            QToolButton:hover {
                background-color: #e0e0e0;
                border-radius: 2px;
            }
        """)
        
        # Conectar sinais
        self.tabCloseRequested.connect(self.close_tab)
        self.tabBar().tabMoved.connect(self.update_add_button_position)
        
        # Criar 3 abas iniciais
        for i in range(3):
            self.add_new_tab()
            
        # Posicionar o botão de adicionar
        self.update_add_button_position()

    def update_add_button_position(self):
        """Atualiza a posição do botão + para ficar após a última aba"""
        tab_bar = self.tabBar()
        if tab_bar.count() > 0:
            last_tab_rect = tab_bar.tabRect(tab_bar.count() - 1)
            x = last_tab_rect.right() + 5
            y = last_tab_rect.top() + (last_tab_rect.height() - self.add_tab_button.height()) // 2
            self.add_tab_button.move(x, y)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_add_button_position()

    def tabInserted(self, index):
        super().tabInserted(index)
        self.update_add_button_position()

    def tabRemoved(self, index):
        super().tabRemoved(index)
        self.update_add_button_position()

    def add_new_tab(self):
        """Adiciona uma nova aba com um QWebEngineView"""
        web_view = QWebEngineView()
        web_view.setUrl(QUrl("https://www.google.com"))
        
        # Configurar a página web com configurações simplificadas
        settings = web_view.settings()
        settings.setAttribute(QWebEngineSettings.PluginsEnabled, False)
        settings.setAttribute(QWebEngineSettings.JavascriptCanOpenWindows, False)
        
        index = self.addTab(web_view, f"Nova aba")
        self.setCurrentIndex(index)
        
        # Atualizar o título quando a página carregar
        web_view.titleChanged.connect(lambda title, view=web_view: 
            self.setTabText(self.indexOf(view), title[:20] + "..." if len(title) > 20 else title))

    def close_tab(self, index):
        """Solicita confirmação e fecha a aba"""
        if self.count() > 1:  # Manter pelo menos uma aba aberta
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Question)
            msg.setText("Deseja fechar esta aba?")
            msg.setWindowTitle("Confirmar fechamento")
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            
            if msg.exec_() == QMessageBox.Yes:
                self.removeTab(index)
        else:
            QMessageBox.warning(self, "Aviso", "Não é possível fechar a última aba.")

if __name__ == '__main__':
    try:
        # Força a liberação de memória antes de iniciar
        import gc
        gc.collect()
        
        # Configura o uso de memória alta prioridade
        app = QApplication(sys.argv)
        app.setStyle('Fusion')
        
        # Configura prioridade de processamento
        import psutil
        process = psutil.Process()
        process.nice(psutil.HIGH_PRIORITY_CLASS)
        
        # Aloca memória para a aplicação
        app.processEvents()
        
        # Monitora uso de memória
        def monitor_memory():
            memory_info = process.memory_info()
            print(f"Uso de memória: {memory_info.rss / 1024 / 1024:.2f} MB")
        
        # Timer para monitorar memória
        timer = QTimer()
        timer.timeout.connect(monitor_memory)
        timer.start(10000)  # Monitora a cada 10 segundos
        
        # Inicia a janela principal
        window = MainWindow()
        window.show()
        
        # Inicia o loop de eventos
        sys.exit(app.exec_())
        
    except Exception as e:
        print(f"Erro na inicialização: {str(e)}")
        sys.exit(1)
    finally:
        # Limpa a memória ao fechar
        gc.collect()
